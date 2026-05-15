from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook, load_workbook

from adbeam_excel_parser.conference_contacts import (
    QUEUE_SHEET_NAME,
    export_icp_contacts_queue,
    read_icp_contacts_summary,
)


class ConferenceContactsTests(unittest.TestCase):
    def test_reads_summary_and_exports_all_rows_to_queue_sheet(self) -> None:
        with TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "icp.xlsx"
            output_path = Path(temp_dir) / "queue.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "База ICP-1"
            worksheet.append([
                "№",
                "Бренд / компания",
                "Сегмент",
                "Пометка из файла",
                "Нужна валидация (?)",
                "Помечено как исключение",
                "Приоритет",
                "Статус обогащения",
                "Юрлицо",
                "ИНН",
                "Выручка, ₽",
                "ОКВЭД / производство",
                "Сайт",
                "ЛПР / CEO",
                "Контакт маркетинга",
                "Email",
                "Телефон",
                "Telegram / соцсети",
                "Источник контактов",
                "Комментарий",
            ])
            worksheet.append([1, "Brand One", "Косметика", None, None, None, "P1", "Не начато"])
            worksheet.append([2, "Brand Two", "Одежда", None, "?", None, "P2", "Не начато"])
            workbook.save(source_path)
            workbook.close()

            summary = read_icp_contacts_summary(source_path)
            exported = export_icp_contacts_queue(source_path, output_path)

            result_workbook = load_workbook(output_path, read_only=False)
            try:
                queue_sheet = result_workbook[QUEUE_SHEET_NAME]
                first_search_cell = queue_sheet.cell(row=2, column=10)

                self.assertEqual(summary.total_rows, 2)
                self.assertEqual(summary.validation_rows, 1)
                self.assertEqual(exported.queue_rows, 2)
                self.assertEqual(queue_sheet.max_row, 3)
                self.assertEqual(queue_sheet.cell(row=2, column=2).value, "Brand One")
                self.assertIsNotNone(first_search_cell.hyperlink)
            finally:
                result_workbook.close()


if __name__ == "__main__":
    unittest.main()
