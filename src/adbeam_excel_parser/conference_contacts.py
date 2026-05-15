from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any, Callable
from urllib.parse import quote_plus

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field


OUTPUT_SUFFIX = "_conference_contacts_queue"
QUEUE_SHEET_NAME = "ICP-1 очередь"
TEAM_SHEET_NAME = "Инструкция ICP-1"

BRAND_HEADER = "Бренд / компания"
SEGMENT_HEADER = "Сегмент"
MARK_HEADER = "Пометка из файла"
VALIDATION_HEADER = "Нужна валидация (?)"
EXCLUSION_HEADER = "Помечено как исключение"
PRIORITY_HEADER = "Приоритет"
STATUS_HEADER = "Статус обогащения"
LEGAL_ENTITY_HEADER = "Юрлицо"
INN_HEADER = "ИНН"
REVENUE_HEADER = "Выручка, ₽"
OKVED_HEADER = "ОКВЭД / производство"
WEBSITE_HEADER = "Сайт"
CEO_HEADER = "ЛПР / CEO"
MARKETING_CONTACT_HEADER = "Контакт маркетинга"
EMAIL_HEADER = "Email"
PHONE_HEADER = "Телефон"
SOCIAL_HEADER = "Telegram / соцсети"
CONTACT_SOURCE_HEADER = "Источник контактов"
COMMENT_HEADER = "Комментарий"

SOURCE_HEADERS = (
    "№",
    BRAND_HEADER,
    SEGMENT_HEADER,
    MARK_HEADER,
    VALIDATION_HEADER,
    EXCLUSION_HEADER,
    PRIORITY_HEADER,
    STATUS_HEADER,
    LEGAL_ENTITY_HEADER,
    INN_HEADER,
    REVENUE_HEADER,
    OKVED_HEADER,
    WEBSITE_HEADER,
    CEO_HEADER,
    MARKETING_CONTACT_HEADER,
    EMAIL_HEADER,
    PHONE_HEADER,
    SOCIAL_HEADER,
    CONTACT_SOURCE_HEADER,
    COMMENT_HEADER,
)

QUEUE_HEADERS = (
    "№",
    BRAND_HEADER,
    SEGMENT_HEADER,
    MARK_HEADER,
    PRIORITY_HEADER,
    VALIDATION_HEADER,
    EXCLUSION_HEADER,
    "Рекомендованный статус",
    "Контур запрос",
    "Поиск сайта",
    "Поиск контактов",
    "Поиск CEO/ЛПР",
    "Поиск маркетинга",
    "Что заполнить в базе",
    "Комментарий оператора",
)

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
LIGHT_GREEN_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
LIGHT_YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
LIGHT_RED_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")
LIGHT_BLUE_FILL = PatternFill(fill_type="solid", fgColor="DDEBF7")
HYPERLINK_FONT = Font(color="0563C1", underline="single")


class IcpContactRowPreview(BaseModel):
    row_index: int
    number: Any | None = None
    brand: str
    segment: str | None = None
    priority: str | None = None
    recommended_status: str


class IcpContactsSummary(BaseModel):
    file_path: str
    output_file_path: str | None = None
    sheet_name: str
    total_rows: int
    queue_rows: int
    validation_rows: int = 0
    excluded_rows: int = 0
    rows_with_inn: int = 0
    rows_with_website: int = 0
    rows_with_email: int = 0
    rows_with_phone: int = 0
    priority_counts: dict[str, int] = Field(default_factory=dict)
    segment_counts: dict[str, int] = Field(default_factory=dict)
    preview: list[IcpContactRowPreview] = Field(default_factory=list)


ProgressCallback = Callable[[int, int, str], None]


def read_icp_contacts_summary(file_path: Path) -> IcpContactsSummary:
    rows, sheet_name = read_icp_rows(file_path)
    return build_summary(
        file_path=file_path,
        sheet_name=sheet_name,
        rows=rows,
        output_file_path=None,
    )


def export_icp_contacts_queue(
    source_file_path: Path,
    output_file_path: Path | None = None,
    progress_callback: ProgressCallback | None = None,
) -> IcpContactsSummary:
    if output_file_path is None:
        output_file_path = build_icp_contacts_output_path(source_file_path)
    if output_file_path.resolve() == source_file_path.resolve():
        raise ValueError("Итоговый файл должен отличаться от исходного Excel")

    rows, sheet_name = read_icp_rows(source_file_path)
    output_file_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(source_file_path)
    try:
        remove_sheet_if_exists(workbook, QUEUE_SHEET_NAME)
        remove_sheet_if_exists(workbook, TEAM_SHEET_NAME)

        queue_sheet = workbook.create_sheet(QUEUE_SHEET_NAME, 0)
        write_queue_sheet(queue_sheet, rows, progress_callback=progress_callback)

        guide_sheet = workbook.create_sheet(TEAM_SHEET_NAME, 1)
        write_team_guide_sheet(guide_sheet)

        workbook.save(output_file_path)
    finally:
        workbook.close()

    return build_summary(
        file_path=source_file_path,
        sheet_name=sheet_name,
        rows=rows,
        output_file_path=output_file_path,
    )


def build_icp_contacts_output_path(source_file_path: Path) -> Path:
    return source_file_path.with_name(f"{source_file_path.stem}{OUTPUT_SUFFIX}{source_file_path.suffix}")


def read_icp_rows(file_path: Path) -> tuple[list[dict[str, Any]], str]:
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    if file_path.suffix.lower() != ".xlsx":
        raise ValueError("Only .xlsx files are supported")

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows_iter = worksheet.iter_rows(values_only=True)
        raw_headers = next(rows_iter, None)
        if raw_headers is None:
            raise ValueError("Excel file is empty")

        headers = [normalize_header(value, index) for index, value in enumerate(raw_headers, start=1)]
        validate_icp_headers(headers)

        rows: list[dict[str, Any]] = []
        for excel_row_index, row_values in enumerate(rows_iter, start=2):
            if not any(value not in (None, "") for value in row_values):
                continue

            row = {header: clean_cell(row_values[index]) if index < len(row_values) else None for index, header in enumerate(headers)}
            row["_excel_row_index"] = excel_row_index
            rows.append(row)

        return rows, worksheet.title
    finally:
        workbook.close()


def validate_icp_headers(headers: list[str]) -> None:
    required_headers = (BRAND_HEADER, SEGMENT_HEADER, PRIORITY_HEADER, STATUS_HEADER)
    missing = [header for header in required_headers if header not in headers]
    if missing:
        raise ValueError("Не найдены обязательные колонки ICP-1: " + ", ".join(missing))


def build_summary(
    file_path: Path,
    sheet_name: str,
    rows: list[dict[str, Any]],
    output_file_path: Path | None,
) -> IcpContactsSummary:
    priority_counts: Counter[str] = Counter()
    segment_counts: Counter[str] = Counter()
    preview: list[IcpContactRowPreview] = []

    validation_rows = 0
    excluded_rows = 0
    rows_with_inn = 0
    rows_with_website = 0
    rows_with_email = 0
    rows_with_phone = 0

    for row in rows:
        priority = as_text(row.get(PRIORITY_HEADER))
        segment = as_text(row.get(SEGMENT_HEADER))
        if priority:
            priority_counts[priority] += 1
        if segment:
            segment_counts[segment] += 1

        if is_truthy_marker(row.get(VALIDATION_HEADER)):
            validation_rows += 1
        if is_truthy_marker(row.get(EXCLUSION_HEADER)):
            excluded_rows += 1
        if as_text(row.get(INN_HEADER)):
            rows_with_inn += 1
        if as_text(row.get(WEBSITE_HEADER)):
            rows_with_website += 1
        if as_text(row.get(EMAIL_HEADER)):
            rows_with_email += 1
        if as_text(row.get(PHONE_HEADER)):
            rows_with_phone += 1

        if len(preview) < 10:
            preview.append(
                IcpContactRowPreview(
                    row_index=int(row.get("_excel_row_index") or 0),
                    number=row.get("№"),
                    brand=as_text(row.get(BRAND_HEADER)),
                    segment=segment or None,
                    priority=priority or None,
                    recommended_status=recommend_status(row),
                )
            )

    return IcpContactsSummary(
        file_path=str(file_path),
        output_file_path=str(output_file_path) if output_file_path is not None else None,
        sheet_name=sheet_name,
        total_rows=len(rows),
        queue_rows=len(rows),
        validation_rows=validation_rows,
        excluded_rows=excluded_rows,
        rows_with_inn=rows_with_inn,
        rows_with_website=rows_with_website,
        rows_with_email=rows_with_email,
        rows_with_phone=rows_with_phone,
        priority_counts=dict(sorted(priority_counts.items())),
        segment_counts=dict(segment_counts.most_common(20)),
        preview=preview,
    )


def write_queue_sheet(
    worksheet,
    rows: list[dict[str, Any]],
    progress_callback: ProgressCallback | None = None,
) -> None:
    worksheet.append(list(QUEUE_HEADERS))
    for column_index in range(1, len(QUEUE_HEADERS) + 1):
        cell = worksheet.cell(row=1, column=column_index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, row in enumerate(rows, start=2):
        brand = as_text(row.get(BRAND_HEADER))
        segment = as_text(row.get(SEGMENT_HEADER))
        kontur_query = build_kontur_query(brand, segment)
        fill_targets = build_fill_targets(row)

        values = [
            row.get("№"),
            brand,
            segment,
            row.get(MARK_HEADER),
            row.get(PRIORITY_HEADER),
            row.get(VALIDATION_HEADER),
            row.get(EXCLUSION_HEADER),
            recommend_status(row),
            kontur_query,
            "найти сайт",
            "найти контакты",
            "найти CEO/ЛПР",
            "найти маркетинг",
            fill_targets,
            "",
        ]
        worksheet.append(values)

        worksheet.cell(row=row_index, column=10).hyperlink = build_search_url(f"{brand} {segment} официальный сайт")
        worksheet.cell(row=row_index, column=11).hyperlink = build_search_url(f"{brand} контакты email телефон официальный сайт")
        worksheet.cell(row=row_index, column=12).hyperlink = build_search_url(f"{brand} генеральный директор основатель")
        worksheet.cell(row=row_index, column=13).hyperlink = build_search_url(f"{brand} директор по маркетингу email")

        for column_index in (10, 11, 12, 13):
            worksheet.cell(row=row_index, column=column_index).font = HYPERLINK_FONT

        row_fill = pick_queue_fill(row)
        for column_index in range(1, len(QUEUE_HEADERS) + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        if progress_callback is not None:
            progress_callback(row_index - 1, len(rows), brand)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    widths = {
        1: 8,
        2: 26,
        3: 34,
        4: 22,
        5: 12,
        6: 16,
        7: 18,
        8: 22,
        9: 44,
        10: 16,
        11: 18,
        12: 18,
        13: 20,
        14: 44,
        15: 34,
    }
    for column_index, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def write_team_guide_sheet(worksheet) -> None:
    rows = [
        ("Шаг", "Что делать", "Результат"),
        ("1", "Искать бренд в Контур.Продажи / Контур.Фокус по колонке 'Контур запрос'.", "Юрлицо, ИНН, выручка, ОКВЭД."),
        ("2", "Если бренд не находится, открывать 'Поиск сайта' и проверять официальный сайт/владельца бренда.", "Корректный сайт и кандидат на юрлицо."),
        ("3", "По ИНН забирать сайт, руководителя, город, выручку и ОКВЭД из Контур/доступных источников.", "ICP-fit и приоритет для кампании."),
        ("4", "Контакты искать в порядке: официальный сайт, карточки компаний, карты, соцсети, пресс/вакансии.", "Email, телефон, Telegram/соцсети, источник."),
        ("5", "Строки с исключением и вопросом обрабатывать отдельно, не грузить в рекламу до проверки.", "Чище база для Директа."),
    ]
    for row in rows:
        worksheet.append(row)

    for column_index in range(1, 4):
        cell = worksheet.cell(row=1, column=column_index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    worksheet.column_dimensions["A"].width = 10
    worksheet.column_dimensions["B"].width = 90
    worksheet.column_dimensions["C"].width = 48
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def remove_sheet_if_exists(workbook, sheet_name: str) -> None:
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])


def build_kontur_query(brand: str, segment: str) -> str:
    parts = [brand]
    if segment:
        parts.append(segment)
    parts.append("производитель")
    return " ".join(part for part in parts if part).strip()


def build_search_url(query: str) -> str:
    return "https://www.google.com/search?q=" + quote_plus(query)


def build_fill_targets(row: dict[str, Any]) -> str:
    missing: list[str] = []
    for header in (
        LEGAL_ENTITY_HEADER,
        INN_HEADER,
        REVENUE_HEADER,
        OKVED_HEADER,
        WEBSITE_HEADER,
        CEO_HEADER,
        MARKETING_CONTACT_HEADER,
        EMAIL_HEADER,
        PHONE_HEADER,
        SOCIAL_HEADER,
        CONTACT_SOURCE_HEADER,
    ):
        if not as_text(row.get(header)):
            missing.append(header)
    return ", ".join(missing)


def recommend_status(row: dict[str, Any]) -> str:
    if is_truthy_marker(row.get(EXCLUSION_HEADER)):
        return "Исключение: проверить вручную"
    if is_truthy_marker(row.get(VALIDATION_HEADER)):
        return "Нужна валидация"
    if not as_text(row.get(INN_HEADER)):
        return "Искать ИНН/юрлицо"
    if not as_text(row.get(WEBSITE_HEADER)):
        return "Искать сайт"
    if not as_text(row.get(EMAIL_HEADER)) and not as_text(row.get(PHONE_HEADER)):
        return "Искать контакты"
    return "Готово к проверке"


def pick_queue_fill(row: dict[str, Any]) -> PatternFill:
    if is_truthy_marker(row.get(EXCLUSION_HEADER)):
        return LIGHT_RED_FILL
    if is_truthy_marker(row.get(VALIDATION_HEADER)):
        return LIGHT_YELLOW_FILL
    if as_text(row.get(INN_HEADER)) or as_text(row.get(WEBSITE_HEADER)):
        return LIGHT_GREEN_FILL
    return LIGHT_BLUE_FILL


def normalize_header(value: Any, column_index: int) -> str:
    text = as_text(value)
    return text or f"column_{column_index}"


def clean_cell(value: Any) -> Any | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip() or None
    return value


def as_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_truthy_marker(value: Any) -> bool:
    text = as_text(value).casefold()
    return text in {"да", "yes", "y", "true", "1", "?", "нужна проверка"}
