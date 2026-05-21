from __future__ import annotations

import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
import os
from typing import Callable
from urllib.parse import parse_qs, unquote, urljoin, urlparse, urlunparse

import httpx
from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field
from selectolax.parser import HTMLParser

from adbeam_excel_parser.conference_contacts import (
    BRAND_HEADER,
    EXCLUSION_HEADER,
    INN_HEADER,
    LEGAL_ENTITY_HEADER,
    PHONE_HEADER,
    PRIORITY_HEADER,
    SEGMENT_HEADER,
    SOCIAL_HEADER,
    VALIDATION_HEADER,
    WEBSITE_HEADER,
    as_text,
    is_truthy_marker,
    read_icp_rows,
)
from adbeam_excel_parser.site_audit import DEFAULT_HEADERS, extract_domain, normalize_url


AUTOSEARCH_SHEET_NAME = "ICP-1 автопоиск"
OUTPUT_SUFFIX = "_icp_autosearch"
MAX_SEARCH_RESULTS = 8
MAX_SITE_PAGES = 4
REQUEST_TIMEOUT_SECONDS = 8.0
DIRECT_DOMAIN_TIMEOUT_SECONDS = 2.0
CONTACT_PAGE_TIMEOUT_SECONDS = 3.0
AUTOSEARCH_WORKERS = 4
DIRECT_DOMAIN_WORKERS = 8
SEARCH_TIMEOUT_SECONDS = 2.0
DUCKDUCKGO_SEARCH_URL = "https://duckduckgo.com/html/"
GOOGLE_SEARCH_URL = "https://www.google.com/search"
YANDEX_SEARCH_URL = "https://yandex.ru/search/"

EMAIL_PATTERN = re.compile(r"(?<![\w.+-])[\w.+-]{2,}@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}(?![\w.+-])")
PHONE_PATTERN = re.compile(r"(?:\+7|8)[\s\-().]*\d{3}[\s\-().]*\d{3}[\s\-().]*\d{2}[\s\-().]*\d{2}")
INN_PATTERN = re.compile(r"(?:инн|ИНН)\D{0,20}(\d{10}|\d{12})")
LEGAL_ENTITY_PATTERN = re.compile(
    r"((?:ООО|ОАО|АО|ЗАО|ПАО|ИП)\s+[\"«]?[A-Za-zА-Яа-яЁё0-9 .,&'()/-]{2,80}[\"»]?)"
)

CONTACT_PATH_TERMS = (
    "contact",
    "contacts",
    "kontakty",
    "kontakti",
    "rekvizit",
    "requisites",
    "about",
    "company",
    "o-kompanii",
    "about-us",
)
CONTACT_TEXT_TERMS = (
    "контакт",
    "реквизит",
    "о компании",
    "contacts",
    "contact",
    "about",
    "company",
)
COMMON_CONTACT_PATHS = (
    "/contacts/",
    "/contacts",
    "/kontakty/",
    "/kontakty",
    "/contact/",
    "/contact",
    "/about/",
    "/about",
    "/company/",
    "/company",
    "/requisites/",
    "/requisites",
    "/rekvizity/",
    "/rekvizity",
)
BLOCKED_DOMAINS = (
    "wildberries.ru",
    "wb.ru",
    "ozon.ru",
    "market.yandex.ru",
    "amazon.com",
    "wikipedia.org",
    "2gis.ru",
    "yandex.ru",
    "google.com",
    "avito.ru",
    "instagram.com",
    "vk.com",
    "t.me",
    "telegram.me",
    "facebook.com",
    "youtube.com",
    "reg.ru",
)
SOCIAL_DOMAINS = (
    "t.me",
    "telegram.me",
    "vk.com",
    "instagram.com",
    "youtube.com",
    "facebook.com",
)
BRAND_DESCRIPTOR_STOPWORDS = {
    "professional",
    "professionnel",
    "laboratories",
    "laboratory",
    "lab",
    "home",
    "cosmetics",
    "cosmetic",
    "selection",
    "makeup",
    "moscow",
    "official",
    "russia",
    "россия",
    "kosmetika",
    "kosmetiki",
    "kosmeticheskaya",
}
BRAND_DOMAIN_ALIASES: tuple[tuple[tuple[str, ...], tuple[str, ...]], ...] = (
    (("trives",), ("trives-spb.ru", "trives-shop.ru")),
    (("черный", "жемчуг"), ("theblackpearl.ru",)),
    (("чёрный", "жемчуг"), ("theblackpearl.ru",)),
    (("юникосметик",), ("estel.beauty", "estel.pro")),
    (("юниккосметик",), ("estel.beauty", "estel.pro")),
    (("чистая", "линия"), ("chistayalinia.ru",)),
    (("невская", "косметика"), ("nevcos.ru",)),
    (("свобода",), ("svobodako.ru",)),
    (("compliment",), ("compliment.su",)),
    (("бабушка", "агаф"), ("1reshenie.ru",)),
    (("рецепты", "бабушки", "агаф"), ("1reshenie.ru",)),
    (("r.o.c.s",), ("rocs.ru",)),
    (("rocs",), ("rocs.ru",)),
)
CYRILLIC_TRANSLIT = str.maketrans(
    {
        "а": "a",
        "б": "b",
        "в": "v",
        "г": "g",
        "д": "d",
        "е": "e",
        "ё": "e",
        "ж": "zh",
        "з": "z",
        "и": "i",
        "й": "y",
        "к": "k",
        "л": "l",
        "м": "m",
        "н": "n",
        "о": "o",
        "п": "p",
        "р": "r",
        "с": "s",
        "т": "t",
        "у": "u",
        "ф": "f",
        "х": "h",
        "ц": "ts",
        "ч": "ch",
        "ш": "sh",
        "щ": "sch",
        "ъ": "",
        "ы": "y",
        "ь": "",
        "э": "e",
        "ю": "yu",
        "я": "ya",
    }
)

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GREEN_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
RED_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")
BLUE_FILL = PatternFill(fill_type="solid", fgColor="DDEBF7")


class SearchCandidate(BaseModel):
    title: str
    url: str
    snippet: str | None = None
    domain: str | None = None
    score: int = 0


class IcpAutoSearchResult(BaseModel):
    row_index: int
    number: str | int | None = None
    brand: str
    segment: str | None = None
    priority: str | None = None
    status: str
    confidence: int = 0
    website: str | None = None
    domain: str | None = None
    title: str | None = None
    emails: list[str] = Field(default_factory=list)
    phones: list[str] = Field(default_factory=list)
    inns: list[str] = Field(default_factory=list)
    legal_entities: list[str] = Field(default_factory=list)
    social_links: list[str] = Field(default_factory=list)
    checked_pages: list[str] = Field(default_factory=list)
    candidates: list[SearchCandidate] = Field(default_factory=list)
    errors: list[str] = Field(default_factory=list)


class IcpAutoSearchSummary(BaseModel):
    file_path: str
    output_file_path: str
    sheet_name: str
    requested_rows: int
    processed_rows: int
    found_websites: int
    found_emails: int
    found_phones: int
    found_inns: int
    high_confidence: int
    results: list[IcpAutoSearchResult] = Field(default_factory=list)


ProgressCallback = Callable[[int, int, str], None]


@dataclass(slots=True)
class FetchedHtml:
    url: str
    status_code: int
    html: str


def run_icp_autosearch(
    source_file_path: Path,
    output_file_path: Path,
    limit: int = 20,
    delay_seconds: float = 1.5,
    only_p1: bool = True,
    progress_callback: ProgressCallback | None = None,
) -> IcpAutoSearchSummary:
    if output_file_path.resolve() == source_file_path.resolve():
        raise ValueError("Итоговый файл должен отличаться от исходного Excel")

    rows, sheet_name = read_icp_rows(source_file_path)
    rows_to_process = select_rows_for_autosearch(rows, limit=limit, only_p1=only_p1)

    results_by_index: dict[int, IcpAutoSearchResult] = {}
    with ThreadPoolExecutor(max_workers=AUTOSEARCH_WORKERS) as executor:
        future_to_index = {}
        for index, row in enumerate(rows_to_process):
            if delay_seconds > 0 and index > 0:
                time.sleep(delay_seconds)
            future = executor.submit(enrich_company_worker, row)
            future_to_index[future] = index

        completed = 0
        for future in as_completed(future_to_index):
            row_index = future_to_index[future]
            result = future.result()
            results_by_index[row_index] = result
            completed += 1
            if progress_callback is not None:
                progress_callback(completed, len(rows_to_process), result.brand)

    results = [results_by_index[index] for index in sorted(results_by_index)]

    output_file_path.parent.mkdir(parents=True, exist_ok=True)
    export_autosearch_results(source_file_path, output_file_path, results)
    return build_autosearch_summary(
        source_file_path=source_file_path,
        output_file_path=output_file_path,
        sheet_name=sheet_name,
        requested_rows=len(rows_to_process),
        results=results,
    )


def enrich_company_worker(row: dict[str, object]) -> IcpAutoSearchResult:
    with httpx.Client(follow_redirects=True, timeout=REQUEST_TIMEOUT_SECONDS, headers=DEFAULT_HEADERS, verify=False) as client:
        return enrich_company_from_open_sources(client, row)


def select_rows_for_autosearch(rows: list[dict[str, object]], limit: int, only_p1: bool) -> list[dict[str, object]]:
    selected: list[dict[str, object]] = []
    for row in rows:
        if only_p1 and as_text(row.get(PRIORITY_HEADER)) != "P1":
            continue
        selected.append(row)
        if limit > 0 and len(selected) >= limit:
            break
    return selected


def enrich_company_from_open_sources(client: httpx.Client, row: dict[str, object]) -> IcpAutoSearchResult:
    brand = as_text(row.get(BRAND_HEADER))
    segment = as_text(row.get(SEGMENT_HEADER))
    errors: list[str] = []
    candidates: list[SearchCandidate] = []

    try:
        candidates = search_company_candidates(client, brand=brand, segment=segment)
    except Exception as exc:
        errors.append(f"search: {exc}")

    best_candidate: SearchCandidate | None = None
    checked_pages: list[FetchedHtml] = []
    extracted = empty_contact_data()
    confidence = 0

    for candidate in candidates[:1]:
        try:
            candidate_pages = crawl_contact_pages(client, candidate.url)
        except Exception as exc:
            errors.append(f"site crawl: {exc}")
            candidate_pages = []

        candidate_extracted = extract_contact_data(candidate_pages)
        candidate_confidence = calculate_confidence(candidate, candidate_extracted)
        if best_candidate is None or candidate_confidence > confidence:
            best_candidate = candidate
            checked_pages = candidate_pages
            extracted = candidate_extracted
            confidence = candidate_confidence

        break

    status = classify_result(best_candidate, extracted, confidence, errors)

    return IcpAutoSearchResult(
        row_index=int(row.get("_excel_row_index") or 0),
        number=row.get("№"),
        brand=brand,
        segment=segment or None,
        priority=as_text(row.get(PRIORITY_HEADER)) or None,
        status=status,
        confidence=confidence,
        website=best_candidate.url if best_candidate is not None else None,
        domain=best_candidate.domain if best_candidate is not None else None,
        title=best_candidate.title if best_candidate is not None else None,
        emails=extracted["emails"],
        phones=extracted["phones"],
        inns=extracted["inns"],
        legal_entities=extracted["legal_entities"],
        social_links=extracted["social_links"],
        checked_pages=[page.url for page in checked_pages],
        candidates=candidates,
        errors=errors,
    )


def search_company_candidates(
    client: httpx.Client,
    brand: str,
    segment: str,
) -> list[SearchCandidate]:
    candidates: list[SearchCandidate] = []
    seen_urls: set[str] = set()

    search_first_candidates = search_engine_first_result_candidates(client, brand, segment, seen_urls)
    if search_first_candidates:
        return search_first_candidates[:MAX_SEARCH_RESULTS]

    for candidate in build_direct_domain_candidates(client, brand, segment):
        candidates.append(candidate)
        seen_urls.add(canonical_url(candidate.url))

    if candidates and candidates[0].score >= 45:
        candidates.sort(key=lambda item: (-item.score, len(item.url)))
        return candidates[:MAX_SEARCH_RESULTS]

    for candidate in search_google_candidates(client, brand, segment, seen_urls):
        candidates.append(candidate)

    candidates.sort(key=lambda item: (-item.score, len(item.url)))
    if candidates and candidates[0].score >= 45:
        return candidates[:MAX_SEARCH_RESULTS]

    for candidate in search_duckduckgo_candidates(client, brand, segment, seen_urls):
        candidates.append(candidate)

    candidates.sort(key=lambda item: (-item.score, len(item.url)))
    return candidates[:MAX_SEARCH_RESULTS]


def search_engine_first_result_candidates(
    client: httpx.Client,
    brand: str,
    segment: str,
    seen_urls: set[str],
) -> list[SearchCandidate]:
    for search_func in (
        search_google_first_result_candidates,
        search_duckduckgo_first_result_candidates,
        search_yandex_first_result_candidates,
    ):
        candidates = search_func(client, brand, segment, seen_urls)
        if candidates:
            return candidates
    return []


def search_google_first_result_candidates(
    client: httpx.Client,
    brand: str,
    segment: str,
    seen_urls: set[str],
) -> list[SearchCandidate]:
    candidates: list[SearchCandidate] = []
    query = build_search_queries(brand, segment)[0]
    try:
        response = client.get(GOOGLE_SEARCH_URL, params={"q": query, "hl": "ru", "gl": "ru", "num": "5"}, timeout=SEARCH_TIMEOUT_SECONDS)
        response.raise_for_status()
    except httpx.HTTPError:
        return candidates

    parser = HTMLParser(response.text)
    for rank, link_node in enumerate(parser.css("div.g a[href], div[data-sokoban-container] a[href], a[href]")):
        url = unwrap_google_url(link_node.attributes.get("href") or "")
        if not url:
            continue
        title = link_node.text(separator=" ", strip=True)
        snippet = extract_nearby_text(link_node)
        if add_google_first_candidate(candidates, seen_urls, brand, segment, url, title, snippet, rank):
            return candidates
    return candidates


def search_duckduckgo_first_result_candidates(
    client: httpx.Client,
    brand: str,
    segment: str,
    seen_urls: set[str],
) -> list[SearchCandidate]:
    candidates: list[SearchCandidate] = []
    query = build_search_queries(brand, segment)[0]
    try:
        response = client.get(DUCKDUCKGO_SEARCH_URL, params={"q": query}, timeout=SEARCH_TIMEOUT_SECONDS)
        response.raise_for_status()
    except httpx.HTTPError:
        return candidates

    parser = HTMLParser(response.text)
    for rank, result_node in enumerate(parser.css(".result")):
        link_node = result_node.css_first("a.result__a")
        if link_node is None:
            continue
        url = unwrap_duckduckgo_url(link_node.attributes.get("href") or "")
        if not url:
            continue
        title = link_node.text(separator=" ", strip=True)
        snippet_node = result_node.css_first(".result__snippet")
        snippet = snippet_node.text(separator=" ", strip=True) if snippet_node else None
        if add_google_first_candidate(candidates, seen_urls, brand, segment, url, title, snippet, rank):
            return candidates
    return candidates


def search_yandex_first_result_candidates(
    client: httpx.Client,
    brand: str,
    segment: str,
    seen_urls: set[str],
) -> list[SearchCandidate]:
    candidates: list[SearchCandidate] = []
    query = build_search_queries(brand, segment)[0]
    try:
        response = client.get(YANDEX_SEARCH_URL, params={"text": query, "lr": "213"}, timeout=SEARCH_TIMEOUT_SECONDS)
        response.raise_for_status()
    except httpx.HTTPError:
        return candidates

    parser = HTMLParser(response.text)
    for rank, result_node in enumerate(parser.css(".serp-item, li")):
        link_node = result_node.css_first("a[href]")
        if link_node is None:
            continue
        url = unwrap_yandex_url(link_node.attributes.get("href") or "")
        if not url:
            continue
        title = link_node.text(separator=" ", strip=True)
        snippet = extract_nearby_text(link_node)
        if add_google_first_candidate(candidates, seen_urls, brand, segment, url, title, snippet, rank):
            return candidates
    return candidates


def search_google_candidates(
    client: httpx.Client,
    brand: str,
    segment: str,
    seen_urls: set[str],
) -> list[SearchCandidate]:
    candidates: list[SearchCandidate] = []

    for query in build_search_queries(brand, segment):
        try:
            response = client.get(GOOGLE_SEARCH_URL, params={"q": query, "hl": "ru"}, timeout=SEARCH_TIMEOUT_SECONDS)
            response.raise_for_status()
        except httpx.HTTPError:
            continue

        parser = HTMLParser(response.text)
        for link_node in parser.css("a[href]"):
            raw_url = link_node.attributes.get("href") or ""
            url = unwrap_google_url(raw_url)
            if not url:
                continue

            title = link_node.text(separator=" ", strip=True)
            snippet = extract_nearby_text(link_node)
            add_search_candidate(
                candidates=candidates,
                seen_urls=seen_urls,
                brand=brand,
                segment=segment,
                url=url,
                title=title,
                snippet=snippet,
            )

    candidates.sort(key=lambda item: (-item.score, len(item.url)))
    return candidates[:MAX_SEARCH_RESULTS]


def search_duckduckgo_candidates(
    client: httpx.Client,
    brand: str,
    segment: str,
    seen_urls: set[str],
) -> list[SearchCandidate]:
    candidates: list[SearchCandidate] = []

    for query in build_search_queries(brand, segment):
        try:
            response = client.get(DUCKDUCKGO_SEARCH_URL, params={"q": query}, timeout=SEARCH_TIMEOUT_SECONDS)
            response.raise_for_status()
        except httpx.HTTPError:
            continue
        parser = HTMLParser(response.text)

        for result_node in parser.css(".result"):
            link_node = result_node.css_first("a.result__a")
            if link_node is None:
                continue

            raw_url = link_node.attributes.get("href") or ""
            url = unwrap_duckduckgo_url(raw_url)
            if not url:
                continue

            title = link_node.text(separator=" ", strip=True)
            snippet_node = result_node.css_first(".result__snippet")
            snippet = snippet_node.text(separator=" ", strip=True) if snippet_node else None
            add_search_candidate(
                candidates=candidates,
                seen_urls=seen_urls,
                brand=brand,
                segment=segment,
                url=url,
                title=title,
                snippet=snippet,
            )

    candidates.sort(key=lambda item: (-item.score, len(item.url)))
    return candidates[:MAX_SEARCH_RESULTS]


def add_search_candidate(
    candidates: list[SearchCandidate],
    seen_urls: set[str],
    brand: str,
    segment: str,
    url: str,
    title: str,
    snippet: str | None,
) -> bool:
    normalized = normalize_url(url)
    if not normalized:
        return False
    candidate_title = title.strip() or extract_domain(normalized) or normalized

    canonical = canonical_url(normalized)
    if canonical in seen_urls:
        return False

    score = score_search_candidate(brand=brand, segment=segment, url=normalized, title=candidate_title, snippet=snippet)
    if score <= 0:
        return False

    seen_urls.add(canonical)
    candidates.append(
        SearchCandidate(
            title=candidate_title,
            url=normalized,
            snippet=snippet,
            domain=extract_domain(normalized),
            score=score,
        )
    )
    return True


def add_google_first_candidate(
    candidates: list[SearchCandidate],
    seen_urls: set[str],
    brand: str,
    segment: str,
    url: str,
    title: str,
    snippet: str | None,
    rank: int,
) -> bool:
    normalized = normalize_url(url)
    if not normalized:
        return False
    candidate_title = title.strip() or extract_domain(normalized) or normalized

    canonical = canonical_url(normalized)
    if canonical in seen_urls:
        return False

    score = score_search_candidate(brand=brand, segment=segment, url=normalized, title=candidate_title, snippet=snippet)
    if score <= 0:
        return False

    seen_urls.add(canonical)
    candidates.append(
        SearchCandidate(
            title=candidate_title,
            url=normalized,
            snippet=snippet,
            domain=extract_domain(normalized),
            score=max(score, 120 - rank),
        )
    )
    return True


def build_search_queries(brand: str, segment: str) -> list[str]:
    del segment
    brand_variants = [brand]
    clean_brand = " ".join(brand_tokens(brand))
    if clean_brand and clean_brand != brand.casefold():
        brand_variants.append(clean_brand)

    queries: list[str] = []
    for variant in unique_values(brand_variants):
        queries.append(f"{variant} официальный сайт")
    return unique_values(queries)[:2]


def build_direct_domain_candidates(client: httpx.Client, brand: str, segment: str) -> list[SearchCandidate]:
    del client
    urls = build_direct_domain_urls(brand)
    candidates: list[SearchCandidate] = []
    seen_urls: set[str] = set()

    if urls:
        first_page = fetch_direct_html(urls[0])
        if first_page is not None:
            add_direct_candidate(candidates, seen_urls, first_page, brand, segment)
            candidates.sort(key=lambda item: (-item.score, len(item.url)))
            if candidates and candidates[0].score >= 45:
                return candidates

    for start in range(1, len(urls), DIRECT_DOMAIN_WORKERS):
        pages: list[FetchedHtml] = []
        batch = urls[start : start + DIRECT_DOMAIN_WORKERS]
        with ThreadPoolExecutor(max_workers=DIRECT_DOMAIN_WORKERS) as executor:
            future_to_url = {executor.submit(fetch_direct_html, url): url for url in batch}
            for future in as_completed(future_to_url):
                page = future.result()
                if page is not None:
                    pages.append(page)

        for page in pages:
            add_direct_candidate(candidates, seen_urls, page, brand, segment)

        candidates.sort(key=lambda item: (-item.score, len(item.url)))
        if candidates and candidates[0].score >= 45:
            break

    return candidates


def add_direct_candidate(
    candidates: list[SearchCandidate],
    seen_urls: set[str],
    page: FetchedHtml,
    brand: str,
    segment: str,
) -> bool:
    canonical = canonical_url(page.url)
    if canonical in seen_urls:
        return False
    seen_urls.add(canonical)

    title = extract_title_from_html(page.html) or urlparse(page.url).netloc
    score = score_search_candidate(brand=brand, segment=segment, url=page.url, title=title, snippet="direct domain guess") + 18
    if is_brand_alias_domain(brand, page.url):
        score = max(score, 68)
    if score < 35:
        return False
    if score <= 0:
        return False

    candidates.append(
        SearchCandidate(
            title=title,
            url=page.url,
            snippet="direct domain guess",
            domain=extract_domain(page.url),
            score=score,
        )
    )
    return True


def is_brand_alias_domain(brand: str, url: str) -> bool:
    domain = extract_domain(url)
    if not domain:
        return False

    aliases = brand_domain_aliases(normalize_brand_for_domain_variants(brand))
    normalized_aliases = {alias.casefold().removeprefix("www.") for alias in aliases}
    return domain.casefold().removeprefix("www.") in normalized_aliases


def fetch_direct_html(url: str) -> FetchedHtml | None:
    try:
        response = httpx.get(
            url,
            follow_redirects=True,
            timeout=DIRECT_DOMAIN_TIMEOUT_SECONDS,
            headers=DEFAULT_HEADERS,
            verify=False,
        )
    except httpx.HTTPError:
        return None

    content_type = response.headers.get("content-type", "").casefold()
    if response.status_code >= 400:
        return None
    if content_type and "html" not in content_type and "text/plain" not in content_type:
        return None
    return FetchedHtml(url=str(response.url), status_code=response.status_code, html=response.text)


def build_direct_domain_urls(brand: str) -> list[str]:
    variants = build_direct_domain_variants(brand)
    url_variants = select_direct_domain_url_variants(variants)

    result: list[str] = []
    for variant in url_variants:
        if len(variant) < 3:
            continue
        if "." in variant:
            result.append(f"https://{variant}/")
            continue
        suffixes = ("com", "ru") if re.search(r"-[a-z]{1,2}$", variant) else ("ru", "com")
        for suffix in suffixes:
            result.append(f"https://{variant}.{suffix}/")
    return result[:32]


def select_direct_domain_url_variants(variants: list[str]) -> list[str]:
    base_variants = [
        variant
        for variant in variants
        if not variant.endswith(("products", "-products"))
    ]
    selected = list(base_variants[:2])
    for variant in base_variants[:2]:
        selected.append(f"{variant}products")
        selected.append(f"{variant}-products")
    selected.extend(base_variants[2:8])
    return unique_values(selected)[:16]


def build_direct_domain_variants(brand: str) -> list[str]:
    lowered = normalize_brand_for_domain_variants(brand)
    without_amp = re.sub(r"[&+]", " ", lowered)
    with_and = re.sub(r"[&+]", " and ", lowered)
    token_lists: list[list[str]] = []
    base_variants: list[str] = brand_domain_aliases(lowered)

    source_values = [without_amp, with_and]
    for value in (without_amp, with_and):
        transliterated = transliterate_cyrillic(value)
        if transliterated != value:
            transliterated_tokens = re.findall(r"[a-z0-9]+", transliterated)
            if transliterated_tokens:
                token_lists.append(transliterated_tokens)
            source_values.append(transliterated)

    for value in unique_values(source_values):
        tokens = re.findall(r"[a-z0-9]+", value)
        if tokens:
            filtered = [token for token in tokens if token not in BRAND_DESCRIPTOR_STOPWORDS and token not in {"and", "the"}]
            if filtered:
                token_lists.append(filtered)
                if len(filtered) == 1 and len(filtered[0]) >= 5:
                    token_lists.append(filtered[:1])
                token_lists.append(filtered[:2])
            token_lists.append(tokens)

    for tokens in token_lists:
        if not tokens:
            continue
        base_variants.extend(build_transliterated_domain_variants(tokens))
        base_variants.append("".join(tokens))
        base_variants.append("-".join(tokens))
        base_variants.extend(build_acronym_variants(tokens))
        base_variants.extend(build_lab_variants(tokens))

    base_variants = unique_values(base_variants)
    variants: list[str] = list(base_variants)
    for variant in base_variants:
        variants.append(variant)
        variants.append(f"{variant}spb")
        variants.append(f"{variant}-spb")
        variants.append(f"{variant}products")
        variants.append(f"{variant}-products")
        variants.append(f"{variant}shop")
        variants.append(f"{variant}-shop")
    return unique_values(variants)


def normalize_brand_for_domain_variants(brand: str) -> str:
    value = brand.casefold()
    value = value.replace("'", "").replace("’", "").replace("`", "")
    return value


def brand_domain_aliases(lowered_brand: str) -> list[str]:
    aliases: list[str] = []
    for required_terms, domains in BRAND_DOMAIN_ALIASES:
        if all(term in lowered_brand for term in required_terms):
            aliases.extend(domains)
    return aliases


def transliterate_cyrillic(value: str) -> str:
    return value.translate(CYRILLIC_TRANSLIT)


def build_transliterated_domain_variants(tokens: list[str]) -> list[str]:
    if not tokens:
        return []

    shortened = [shorten_domain_token(token) for token in tokens]
    has_shortcut = (
        shortened != tokens
        or compound_initial_code(tokens[0]) != (tokens[0][0] if tokens[0] else "")
        or any(token.startswith("veno") for token in tokens)
    )
    if not has_shortcut:
        return []

    variants: list[str] = []
    for token in tokens:
        if token.startswith("veno"):
            variants.append("venoshop")
            variants.append("veno-shop")

    if len(tokens) >= 2:
        first_short = shorten_domain_token(tokens[0])
        rest_initials = "".join(token[0] for token in tokens[1:] if token)
        if first_short != tokens[0] and rest_initials:
            variants.append(f"{first_short}-{rest_initials}")
            variants.append(f"{first_short}{rest_initials}")

        first_code = compound_initial_code(tokens[0])
        last_short = shorten_domain_token(tokens[-1])
        variants.append(f"{first_code}{last_short}")
        variants.append(f"{first_code}-{last_short}")

        if rest_initials:
            variants.append(f"{first_short}{rest_initials}")
            variants.append(f"{first_short}-{rest_initials}")

    if shortened != tokens:
        variants.append("".join(shortened))
        variants.append("-".join(shortened))

    joined = "".join(tokens)
    dashed = "-".join(tokens)
    for value in (joined, dashed):
        if "iya" in value:
            variants.append(value.replace("iya", "ia"))

    return unique_values([variant for variant in variants if len(variant) >= 3])


def shorten_domain_token(token: str) -> str:
    if token.startswith("kosmet"):
        return "cosm"
    if token.startswith("master"):
        return "master"
    if token.startswith("parfyum") or token.startswith("parfum"):
        return "parfum"
    return token


def compound_initial_code(token: str) -> str:
    if token.startswith("krasnopol") or token.startswith("krasno"):
        return "kp"
    return token[0] if token else ""


def build_acronym_variants(tokens: list[str]) -> list[str]:
    tokens = normalize_acronym_tokens(tokens)
    if len(tokens) < 2:
        return []

    acronym = "".join(token[0] for token in tokens if token)
    variants = [acronym]
    last_token = tokens[-1]
    if len(last_token) >= 4:
        variants.append(f"{acronym}{last_token}")

    # Don't Touch My Skin -> dtmskin, d-t-m-skin
    if len(tokens) >= 3 and len(last_token) >= 4:
        variants.append("".join(token[0] for token in tokens[:-1]) + last_token)
        variants.append("-".join([*(token[0] for token in tokens[:-1]), last_token]))
    return variants


def normalize_acronym_tokens(tokens: list[str]) -> list[str]:
    result: list[str] = []
    index = 0
    while index < len(tokens):
        if index + 1 < len(tokens) and tokens[index] == "don" and tokens[index + 1] == "t":
            result.append("dont")
            index += 2
            continue
        if tokens[index]:
            result.append(tokens[index])
        index += 1
    return result


def build_lab_variants(tokens: list[str]) -> list[str]:
    variants: list[str] = []
    for token in tokens:
        if token.endswith("lab") and len(token) > 4:
            prefix = token[:-3]
            variants.append(f"{prefix}-lab")
            variants.append(f"{prefix}lab")
    return variants


def unwrap_duckduckgo_url(raw_url: str) -> str | None:
    if not raw_url:
        return None

    if raw_url.startswith("//"):
        raw_url = "https:" + raw_url

    parsed = urlparse(raw_url)
    if "duckduckgo.com" in parsed.netloc and parsed.path.startswith("/l/"):
        target = parse_qs(parsed.query).get("uddg", [""])[0]
        return unquote(target) if target else None

    if parsed.scheme in {"http", "https"} and parsed.netloc:
        return raw_url

    return None


def unwrap_google_url(raw_url: str) -> str | None:
    if not raw_url:
        return None

    if raw_url.startswith("//"):
        raw_url = "https:" + raw_url
    elif raw_url.startswith("/"):
        raw_url = urljoin(GOOGLE_SEARCH_URL, raw_url)

    parsed = urlparse(raw_url)
    domain = extract_domain(raw_url)
    if domain == "google.com":
        if parsed.path == "/url":
            query = parse_qs(parsed.query)
            target = query.get("q", [""])[0] or query.get("url", [""])[0]
            return unquote(target) if target else None
        return None

    if parsed.scheme in {"http", "https"} and parsed.netloc:
        return raw_url

    return None


def unwrap_yandex_url(raw_url: str) -> str | None:
    if not raw_url:
        return None

    if raw_url.startswith("//"):
        raw_url = "https:" + raw_url
    elif raw_url.startswith("/"):
        raw_url = urljoin(YANDEX_SEARCH_URL, raw_url)

    parsed = urlparse(raw_url)
    domain = extract_domain(raw_url)
    if domain in {"yandex.ru", "ya.ru"}:
        query = parse_qs(parsed.query)
        target = query.get("url", [""])[0] or query.get("to", [""])[0]
        return unquote(target) if target else None

    if parsed.scheme in {"http", "https"} and parsed.netloc:
        return raw_url

    return None


def extract_nearby_text(node) -> str | None:
    best = node.text(separator=" ", strip=True)
    current = node
    for _ in range(4):
        current = current.parent
        if current is None:
            break
        text = current.text(separator=" ", strip=True)
        if len(text) > len(best):
            best = text
        if len(best) >= 400:
            break

    cleaned = re.sub(r"\s+", " ", best).strip()
    return cleaned[:800] if cleaned else None


def score_search_candidate(brand: str, segment: str, url: str, title: str, snippet: str | None) -> int:
    domain = extract_domain(url)
    if not domain or domain.casefold() in BLOCKED_DOMAINS:
        return -100

    parsed = urlparse(url)
    haystack = f"{parsed.netloc} {parsed.path} {title} {snippet or ''}".casefold()
    if is_known_false_positive(brand, segment, domain, haystack):
        return -100
    tokens = brand_tokens(brand)

    score = 0
    token_matches = 0

    for token in tokens:
        if token in haystack:
            score += 12
            token_matches += 1
        if token in parsed.netloc.casefold():
            score += 10
            token_matches += 1

    if tokens and token_matches == 0:
        return -10

    if any(term in haystack for term in ("официаль", "official", "контакт", "contact")):
        score += 8
    score += score_segment_fit(segment, haystack)
    score += score_domain_zone(parsed.netloc)
    if parsed.path in {"", "/"}:
        score += 5
    if len(parsed.path.strip("/").split("/")) <= 1:
        score += 3

    if any(term in domain.casefold() for term in ("rusprofile", "list-org", "spark", "sbis", "zachestnyibiznes")):
        score -= 20
    if any(term in haystack for term in ("отзывы", "review", "маркетплейс", "купить на")):
        score -= 8

    return score


def is_known_false_positive(brand: str, segment: str, domain: str, haystack: str) -> bool:
    brand_lower = brand.casefold()
    domain_lower = domain.casefold().removeprefix("www.")

    if "doctor wax" in brand_lower and domain_lower == "doctor.ru":
        return True
    if not is_food_segment(segment) and ("magnit" in brand_lower or "магнит" in brand_lower):
        return True
    if domain_lower == "reg.ru":
        return True
    if "domain/shop" in haystack or "домен прода" in haystack:
        return True

    return False


def score_domain_zone(netloc: str) -> int:
    host = netloc.casefold().removeprefix("www.")
    if host.endswith(".ru") or host.endswith(".рф") or host.endswith(".xn--p1ai"):
        return 8
    if host.endswith((".fr", ".ua", ".by", ".kz", ".al", ".pl", ".de")):
        return -10
    return 0


def score_segment_fit(segment: str, haystack: str) -> int:
    segment_lower = segment.casefold()
    if any(term in segment_lower for term in ("космет", "парфюмер", "гигиен")):
        if any(term in haystack for term in ("космет", "cosmetic", "beauty", "уход", "skin", "makeup", "макияж", "парфюм")):
            return 20
        if any(term in haystack for term in ("игр", "game", "gaming", "разработки игр", "plastic surgery")):
            return -45
    if any(term in segment_lower for term in ("одеж", "обув", "текстил")):
        if any(term in haystack for term in ("одеж", "обув", "fashion", "wear", "shoes", "textile")):
            return 15
    return 0


def is_food_segment(segment: str) -> bool:
    segment_lower = segment.casefold()
    return any(term in segment_lower for term in ("продукт", "еда", "питани", "food", "grocery"))


def brand_tokens(brand: str) -> list[str]:
    raw_tokens = re.findall(r"[a-zA-Zа-яА-ЯёЁ0-9]+", brand.casefold())
    return [token for token in raw_tokens if len(token) >= 3]


def crawl_contact_pages(client: httpx.Client, site_url: str) -> list[FetchedHtml]:
    normalized = normalize_url(site_url)
    if not normalized:
        return []

    pages: list[FetchedHtml] = []
    seen: set[str] = set()
    entry_page = fetch_html(client, normalized)
    if entry_page is None:
        return []

    pages.append(entry_page)
    seen.add(canonical_url(entry_page.url))
    for candidate_url in build_contact_page_candidates(entry_page):
        if len(pages) >= MAX_SITE_PAGES:
            break
        canonical = canonical_url(candidate_url)
        if canonical in seen:
            continue
        seen.add(canonical)
        page = fetch_html(client, candidate_url, timeout_seconds=CONTACT_PAGE_TIMEOUT_SECONDS)
        if page is not None:
            pages.append(page)

    return pages


def fetch_html(client: httpx.Client, url: str, timeout_seconds: float | None = None) -> FetchedHtml | None:
    try:
        response = client.get(url, timeout=timeout_seconds or REQUEST_TIMEOUT_SECONDS)
    except httpx.HTTPError:
        return None
    content_type = response.headers.get("content-type", "").casefold()
    if response.status_code >= 400:
        return None
    if content_type and "html" not in content_type and "text/plain" not in content_type:
        return None
    return FetchedHtml(url=str(response.url), status_code=response.status_code, html=response.text)


def build_contact_page_candidates(entry_page: FetchedHtml) -> list[str]:
    parser = HTMLParser(entry_page.html)
    scored: dict[str, int] = {}
    domain = extract_domain(entry_page.url)

    for node in parser.css("a[href]"):
        href = (node.attributes.get("href") or "").strip()
        text = node.text(separator=" ", strip=True)
        candidate = normalize_site_link(entry_page.url, href)
        if not candidate or extract_domain(candidate) != domain:
            continue

        score = score_contact_link(candidate, text)
        if score > 0:
            scored[candidate] = max(score, scored.get(candidate, 0))

    parsed = urlparse(entry_page.url)
    origin = urlunparse((parsed.scheme, parsed.netloc, "", "", "", ""))
    for index, path in enumerate(COMMON_CONTACT_PATHS):
        candidate = urljoin(origin, path)
        scored.setdefault(candidate, 50 - index)

    return [
        url
        for url, _score in sorted(
            scored.items(),
            key=lambda item: (-item[1], len(urlparse(item[0]).path), item[0]),
        )
    ][:MAX_SITE_PAGES]


def normalize_site_link(base_url: str, href: str) -> str | None:
    if not href:
        return None

    lowered = href.casefold()
    if lowered.startswith(("mailto:", "tel:", "javascript:", "#")):
        return None

    url = urljoin(base_url, href)
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        return None
    if any(parsed.path.casefold().endswith(extension) for extension in (".pdf", ".jpg", ".png", ".zip", ".rar", ".xlsx")):
        return None
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path or "/", "", parsed.query or "", ""))


def score_contact_link(url: str, text: str) -> int:
    value = f"{url} {text}".casefold()
    score = 0
    if any(term in value for term in CONTACT_PATH_TERMS):
        score += 12
    if any(term in value for term in CONTACT_TEXT_TERMS):
        score += 12
    if any(term in value for term in ("privacy", "policy", "cart", "basket", "catalog", "product")):
        score -= 8
    return score


def extract_contact_data(pages: list[FetchedHtml]) -> dict[str, list[str]]:
    emails: list[str] = []
    phones: list[str] = []
    inns: list[str] = []
    legal_entities: list[str] = []
    social_links: list[str] = []

    for page in pages:
        parser = HTMLParser(page.html)
        text = parser.body.text(separator=" ", strip=True) if parser.body else parser.text(separator=" ", strip=True)
        html = page.html

        emails.extend(match.group(0) for match in EMAIL_PATTERN.finditer(text))
        emails.extend(match.group(0) for match in EMAIL_PATTERN.finditer(html))
        phones.extend(normalize_phone(match.group(0)) for match in PHONE_PATTERN.finditer(text))
        inns.extend(match.group(1) for match in INN_PATTERN.finditer(text))
        legal_entities.extend(clean_legal_entity(match.group(1)) for match in LEGAL_ENTITY_PATTERN.finditer(text))
        social_links.extend(extract_social_links(parser, page.url))

    return {
        "emails": unique_values([email for email in emails if is_useful_email(email)])[:8],
        "phones": unique_values([phone for phone in phones if phone])[:8],
        "inns": unique_values(inns)[:5],
        "legal_entities": unique_values([entity for entity in legal_entities if entity])[:5],
        "social_links": unique_values(social_links)[:8],
    }


def empty_contact_data() -> dict[str, list[str]]:
    return {
        "emails": [],
        "phones": [],
        "inns": [],
        "legal_entities": [],
        "social_links": [],
    }


def extract_title_from_html(html: str) -> str | None:
    if not html:
        return None
    parser = HTMLParser(html)
    title_node = parser.css_first("title")
    if title_node is None:
        return None
    title = title_node.text(separator=" ", strip=True)
    return title or None


def extract_social_links(parser: HTMLParser, base_url: str) -> list[str]:
    links: list[str] = []
    for node in parser.css("a[href]"):
        href = node.attributes.get("href") or ""
        if href.startswith("//"):
            href = "https:" + href
        if href.startswith("/"):
            href = urljoin(base_url, href)
        domain = extract_domain(href)
        if domain and domain.casefold() in SOCIAL_DOMAINS:
            links.append(href)
    return links


def calculate_confidence(candidate: SearchCandidate | None, extracted: dict[str, list[str]]) -> int:
    if candidate is None:
        return 0
    score = min(max(candidate.score, 0), 70)
    if extracted["emails"]:
        score += 12
    if extracted["phones"]:
        score += 10
    if extracted["inns"]:
        score += 8
    return min(score, 100)


def classify_result(
    candidate: SearchCandidate | None,
    extracted: dict[str, list[str]],
    confidence: int,
    errors: list[str],
) -> str:
    del confidence, errors
    if candidate is None:
        return "NOT_FOUND"
    if extracted["emails"]:
        return "FOUND_CONTACTS"
    return "FOUND_SITE_CHECK_CONTACTS"


def export_autosearch_results(source_file_path: Path, output_file_path: Path, results: list[IcpAutoSearchResult]) -> None:
    workbook = load_workbook(source_file_path)
    temp_output_path = output_file_path.with_name(f"{output_file_path.stem}.tmp{output_file_path.suffix}")
    try:
        if AUTOSEARCH_SHEET_NAME in workbook.sheetnames:
            workbook.remove(workbook[AUTOSEARCH_SHEET_NAME])
        sheet = workbook.create_sheet(AUTOSEARCH_SHEET_NAME, 0)
        write_autosearch_sheet(sheet, results)
        workbook.save(temp_output_path)
    finally:
        workbook.close()

    validate_xlsx_file(temp_output_path)
    os.replace(temp_output_path, output_file_path)


def write_autosearch_sheet(sheet, results: list[IcpAutoSearchResult]) -> None:
    headers = [
        "№",
        BRAND_HEADER,
        SEGMENT_HEADER,
        PRIORITY_HEADER,
        "Авто статус",
        "Уверенность",
        WEBSITE_HEADER,
        "Домен",
        "Title",
        "Email",
        PHONE_HEADER,
        INN_HEADER,
        LEGAL_ENTITY_HEADER,
        SOCIAL_HEADER,
        "Проверенные страницы",
        "Кандидаты",
        "Ошибки",
    ]
    sheet.append(headers)
    for column_index in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=column_index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, result in enumerate(results, start=2):
        values = [
            result.number,
            result.brand,
            result.segment,
            result.priority,
            result.status,
            result.confidence,
            result.website,
            result.domain,
            result.title,
            join_values(result.emails),
            join_values(result.phones),
            join_values(result.inns),
            join_values(result.legal_entities),
            join_values(result.social_links),
            join_values(result.checked_pages),
            join_values([f"{candidate.score}: {candidate.url}" for candidate in result.candidates]),
            join_values(result.errors),
        ]
        sheet.append([safe_excel_value(value) for value in values])
        fill = pick_result_fill(result)
        for column_index in range(1, len(headers) + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if result.website:
            sheet.cell(row=row_index, column=7).hyperlink = result.website
            sheet.cell(row=row_index, column=7).font = Font(color="0563C1", underline="single")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {
        1: 8,
        2: 26,
        3: 30,
        4: 12,
        5: 24,
        6: 12,
        7: 34,
        8: 22,
        9: 36,
        10: 34,
        11: 24,
        12: 18,
        13: 34,
        14: 34,
        15: 52,
        16: 52,
        17: 36,
    }
    for column_index, width in widths.items():
        sheet.column_dimensions[get_column_letter(column_index)].width = width


def build_autosearch_summary(
    source_file_path: Path,
    output_file_path: Path,
    sheet_name: str,
    requested_rows: int,
    results: list[IcpAutoSearchResult],
) -> IcpAutoSearchSummary:
    return IcpAutoSearchSummary(
        file_path=str(source_file_path),
        output_file_path=str(output_file_path),
        sheet_name=sheet_name,
        requested_rows=requested_rows,
        processed_rows=len(results),
        found_websites=sum(1 for result in results if result.website),
        found_emails=sum(1 for result in results if result.emails),
        found_phones=sum(1 for result in results if result.phones),
        found_inns=sum(1 for result in results if result.inns),
        high_confidence=sum(1 for result in results if result.confidence >= 70),
        results=results,
    )


def validate_xlsx_file(file_path: Path) -> None:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    workbook.close()


def safe_excel_value(value):
    if isinstance(value, str):
        return sanitize_excel_text(value)
    return value


def sanitize_excel_text(value: str) -> str:
    cleaned = ILLEGAL_CHARACTERS_RE.sub("", value)
    cleaned = "".join(
        character
        for character in cleaned
        if character in "\t\n\r" or ord(character) >= 32
    )
    if len(cleaned) > 32767:
        return cleaned[:32767]
    return cleaned


def pick_result_fill(result: IcpAutoSearchResult) -> PatternFill:
    if result.emails:
        return GREEN_FILL
    if result.website:
        return YELLOW_FILL
    if result.status == "NOT_FOUND":
        return RED_FILL
    return BLUE_FILL


def canonical_url(url: str) -> str:
    parsed = urlparse(url)
    path = parsed.path.rstrip("/") or "/"
    return urlunparse((parsed.scheme.casefold(), parsed.netloc.casefold(), path, "", parsed.query or "", ""))


def normalize_phone(value: str) -> str:
    digits = re.sub(r"\D+", "", value)
    if len(digits) == 11 and digits.startswith("8"):
        digits = "7" + digits[1:]
    if len(digits) == 11 and digits.startswith("7"):
        return f"+7 {digits[1:4]} {digits[4:7]}-{digits[7:9]}-{digits[9:11]}"
    return value.strip()


def clean_legal_entity(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip(" .,;:-")


def is_useful_email(value: str) -> bool:
    lowered = value.casefold()
    local_part = lowered.split("@", maxsplit=1)[0]
    domain = lowered.split("@", maxsplit=1)[1] if "@" in lowered else ""
    if local_part in {"test", "example", "noreply", "no-reply"}:
        return False
    if local_part in {"email", "your-email", "your_email"}:
        return False
    if domain.startswith("example.") or domain in {"example.ru", "example.com"}:
        return False
    if "sentry" in domain or "ingest." in domain:
        return False
    return not any(lowered.endswith(extension) for extension in (".png", ".jpg", ".jpeg", ".gif", ".webp"))


def unique_values(values: list[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        normalized = value.strip()
        if not normalized:
            continue
        key = normalized.casefold()
        if key in seen:
            continue
        seen.add(key)
        result.append(normalized)
    return result


def join_values(values: list[str]) -> str:
    return "; ".join(values)


def build_icp_autosearch_output_path(source_file_path: Path) -> Path:
    return source_file_path.with_name(f"{source_file_path.stem}{OUTPUT_SUFFIX}{source_file_path.suffix}")
