from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
import zipfile

from openpyxl import Workbook, load_workbook


EXCEL_SUFFIXES = {".xlsx", ".xlsm"}


def merge_excel_files_to_tabs(
    source_path: Path,
    output_xlsx_path: Path,
    output_zip_path: Path | None = None,
    create_zip: bool = True,
) -> dict[str, str | int | None]:
    source_path = Path(source_path)
    output_xlsx_path = Path(output_xlsx_path)
    output_zip_path = Path(output_zip_path) if output_zip_path is not None else None

    if not source_path.exists():
        raise FileNotFoundError(f"Источник не найден: {source_path}")

    if source_path.is_dir():
        excel_files = _discover_excel_files(
            source_root=source_path,
            excluded_paths={_safe_resolve(output_xlsx_path)},
        )
        return _merge_discovered_excel_files(
            source_path=source_path,
            excel_files=excel_files,
            output_xlsx_path=output_xlsx_path,
            output_zip_path=output_zip_path,
            create_zip=create_zip,
        )

    if source_path.is_file():
        if not zipfile.is_zipfile(source_path):
            raise ValueError(f"Источник должен быть ZIP-архивом или папкой: {source_path}")

        with TemporaryDirectory(prefix="wb_excel_merge_") as temp_dir:
            with zipfile.ZipFile(source_path, "r") as archive:
                archive.extractall(temp_dir)

            excel_files = _discover_excel_files(
                source_root=Path(temp_dir),
                excluded_paths=set(),
            )
            return _merge_discovered_excel_files(
                source_path=source_path,
                excel_files=excel_files,
                output_xlsx_path=output_xlsx_path,
                output_zip_path=output_zip_path,
                create_zip=create_zip,
            )

    raise ValueError(f"Источник должен быть ZIP-архивом или папкой: {source_path}")


def _discover_excel_files(source_root: Path, excluded_paths: set[Path]) -> list[Path]:
    excel_files: list[Path] = []
    for path in source_root.rglob("*"):
        if not path.is_file():
            continue
        if path.name.startswith("~$"):
            continue
        if path.suffix.casefold() not in EXCEL_SUFFIXES:
            continue
        if _safe_resolve(path) in excluded_paths:
            continue
        excel_files.append(path)

    return sorted(excel_files, key=lambda path: (path.name.casefold(), str(path).casefold()))


def _merge_discovered_excel_files(
    source_path: Path,
    excel_files: list[Path],
    output_xlsx_path: Path,
    output_zip_path: Path | None,
    create_zip: bool,
) -> dict[str, str | int | None]:
    if not excel_files:
        raise ValueError(f"В источнике не найдено Excel-файлов (.xlsx, .xlsm): {source_path}")

    output_xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    try:
        for index, excel_path in enumerate(excel_files, start=1):
            sheet = workbook.create_sheet(title=str(index))
            source_workbook = _load_source_workbook(excel_path)
            try:
                source_sheet = source_workbook.worksheets[0]
                for row in source_sheet.iter_rows(values_only=True):
                    sheet.append(list(row))
            finally:
                try:
                    source_workbook.close()
                except Exception:
                    pass

        try:
            workbook.save(output_xlsx_path)
        except Exception as exc:
            raise RuntimeError(f"Не удалось сохранить итоговый Excel: {output_xlsx_path}") from exc
    finally:
        try:
            workbook.close()
        except Exception:
            pass

    created_zip_path: Path | None = None
    if create_zip:
        created_zip_path = output_zip_path or output_xlsx_path.with_suffix(".zip")
        created_zip_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            with zipfile.ZipFile(created_zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
                archive.write(output_xlsx_path, arcname=output_xlsx_path.name)
        except Exception as exc:
            raise RuntimeError(f"Не удалось создать ZIP-архив: {created_zip_path}") from exc

    return {
        "source_path": str(source_path),
        "files_found": len(excel_files),
        "sheets_created": len(excel_files),
        "output_xlsx_path": str(output_xlsx_path),
        "output_zip_path": str(created_zip_path) if created_zip_path is not None else None,
    }


def _load_source_workbook(excel_path: Path):
    try:
        return load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as exc:
        raise RuntimeError(f"Не удалось открыть Excel-файл: {excel_path}") from exc


def _safe_resolve(path: Path) -> Path:
    try:
        return path.resolve()
    except Exception:
        return path.absolute()
