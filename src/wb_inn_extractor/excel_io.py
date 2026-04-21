from __future__ import annotations

from pathlib import Path
from collections import Counter
from datetime import datetime
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import AnalyzeSummary, ResearchRow, SheetAnalyzeSummary


REQUIRED_HEADERS = {
    "product_name": "Товар",
    "article": "Артикул",
    "brand": "Бренд",
    "seller": "Продавец",
}

HEADER_SCAN_LIMIT = 20
PREVIEW_LIMIT = 5
COMPASS_INN_HEADER_CANDIDATES = [
    "ИНН",
    "ИНН ЮЛ",
    "ИНН ЮЛ/ИП",
    "ИНН компании",
    "ИНН контрагента",
]

INN_REGISTRY_HEADERS = [
    "first_category",
    "seller_name_raw",
    "seller_display_name",
    "entity_type",
    "inn",
    "ogrn",
    "ogrnip",
    "seller_url",
    "wb_candidate_url",
    "parse_status",
    "note",
    "seen_count",
    "first_seen_at",
    "last_seen_at",
    "first_source_batch",
    "last_source_batch",
]

INN_IMPORT_HEADERS = [
    "imported_at",
    "source_batch_file",
    "category",
    "rows_total",
    "rows_with_inn",
    "unique_inn_in_file",
    "new_inn_added",
    "already_known",
    "duplicate_in_file",
    "skipped_without_inn",
]

SELLER_HISTORY_HEADERS = [
    "seller_key",
    "first_seller_name_raw",
    "last_seller_name_raw",
    "seller_display_name",
    "seller_url",
    "entity_type",
    "inn",
    "ogrn",
    "ogrnip",
    "first_parse_status",
    "last_parse_status",
    "first_note",
    "last_note",
    "skip_recommended",
    "seen_count",
    "first_seen_at",
    "last_seen_at",
    "first_source_batch",
    "last_source_batch",
    "first_category",
    "last_category",
]

SELLER_HISTORY_IMPORT_HEADERS = [
    "imported_at",
    "source_batch_file",
    "category",
    "rows_total",
    "rows_with_seller",
    "unique_seller_in_file",
    "new_seller_added",
    "already_known",
    "duplicate_in_file",
    "skipped_without_seller",
    "skip_recommended_in_file",
]

SELLER_HISTORY_SKIP_STATUSES = {
    "SUCCESS",
    "PARTIAL_SUCCESS",
    "PAGE_OPENED_NO_REQUISITES",
    "Нет реквизитов на странице",
    "БЕЛОРУСЬ",
    "КАЗАХСТАН",
    "SKIPPED_KNOWN_SELLER",
    "SKIPPED_VIEWED_SELLER",
}

BATCH_RESULTS_REQUIRED_HEADERS = {
    "inn",
    "seller_name_raw",
    "wb_candidate_url",
    "parse_status",
}


def load_active_sheet(input_path: Path):
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    if not workbook.sheetnames:
        raise ValueError("В книге нет листов")
    return workbook[workbook.sheetnames[0]]


def analyze_workbook(input_path: Path, selected_sheets: list[str] | None = None) -> AnalyzeSummary:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    if not workbook.sheetnames:
        raise ValueError("В книге нет листов")

    normalized_selection = _normalize_sheet_selection(workbook.sheetnames, selected_sheets)
    summaries: list[SheetAnalyzeSummary] = []
    valid_sheet_names: list[str] = []
    skipped_sheet_names: list[str] = []

    for sheet_name in normalized_selection:
        sheet = workbook[sheet_name]
        summary = _analyze_sheet(sheet)
        summaries.append(summary)
        if summary.status == "VALID":
            valid_sheet_names.append(sheet_name)
        else:
            skipped_sheet_names.append(sheet_name)

    first_valid = next((item for item in summaries if item.status == "VALID"), None)
    return AnalyzeSummary(
        input_path=input_path,
        workbook_sheet_count=len(workbook.sheetnames),
        selected_sheet_names=normalized_selection,
        skipped_sheet_names=skipped_sheet_names,
        valid_sheet_names=valid_sheet_names,
        sheet_name=first_valid.sheet_name if first_valid else None,
        total_rows=first_valid.data_rows_count if first_valid else 0,
        total_columns=first_valid.total_columns if first_valid else 0,
        headers=first_valid.headers if first_valid else [],
        has_seller_name=first_valid.has_seller_name if first_valid else False,
        has_brand=first_valid.has_brand if first_valid else False,
        has_wb_url=first_valid.has_wb_url if first_valid else False,
        has_article=first_valid.has_article if first_valid else False,
        candidate_key_field=first_valid.candidate_key_field if first_valid else None,
        preview_rows=first_valid.preview_rows if first_valid else [],
        sheets=summaries,
    )


def build_wb_candidate_url(nm_id: int | None) -> str | None:
    if nm_id is None:
        return None
    return f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx"


def extract_research_rows(
    input_path: Path,
    limit: int | None = None,
    selected_sheets: list[str] | None = None,
    excluded_seller_keys: set[str] | None = None,
) -> list[ResearchRow]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet_names = _normalize_sheet_selection(workbook.sheetnames, selected_sheets)

    result: list[ResearchRow] = []
    seen_sellers: set[str] = set()

    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        sheet_meta = _prepare_sheet_meta(sheet)
        if sheet_meta is None:
            continue

        for excel_row_index, row in enumerate(
            sheet.iter_rows(min_row=sheet_meta["header_row_index"] + 1, values_only=True),
            start=sheet_meta["header_row_index"] + 1,
        ):
            row_values = list(row)
            if _is_effectively_empty_row(row_values):
                continue

            nm_id = _coerce_wb_nm_id(_get_cell(row_values, sheet_meta["header_index"], REQUIRED_HEADERS["article"]))
            brand_raw = _coerce_to_string(_get_cell(row_values, sheet_meta["header_index"], REQUIRED_HEADERS["brand"]))
            seller_raw = _coerce_to_string(_get_cell(row_values, sheet_meta["header_index"], REQUIRED_HEADERS["seller"]))
            product_name = _coerce_to_string(_get_cell(row_values, sheet_meta["header_index"], REQUIRED_HEADERS["product_name"]))

            seller_key = _normalize_key_part(seller_raw)
            if seller_key:
                if seller_key in seen_sellers:
                    continue
                if excluded_seller_keys and seller_key in excluded_seller_keys:
                    continue
                seen_sellers.add(seller_key)

            research_row = ResearchRow(
                source_sheet=sheet.title,
                source_row_index=excel_row_index,
                product_name=product_name,
                wb_nm_id=nm_id,
                brand=brand_raw,
                seller_name_raw=seller_raw,
                wb_candidate_url=build_wb_candidate_url(nm_id),
                parse_status="READY_FOR_RESEARCH" if nm_id else "NO_NM_ID",
                parse_note=None if nm_id else "В строке нет корректного Артикул/nmID",
            )
            result.append(research_row)
            if limit is not None and len(result) >= limit:
                return result
    return result


def summarize_research_rows_by_sheet(rows: list[ResearchRow]) -> dict[str, int]:
    counter = Counter(row.source_sheet for row in rows)
    return dict(sorted(counter.items(), key=lambda item: (item[0].casefold(), item[0])))


def save_research_sample(output_path: Path, rows: list[ResearchRow]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "research_sample"
    headers = [
        "source_sheet",
        "source_row_index",
        "product_name",
        "wb_nm_id",
        "brand",
        "seller_name_raw",
        "wb_candidate_url",
        "parse_status",
        "parse_note",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append([
            row.source_sheet,
            row.source_row_index,
            row.product_name,
            row.wb_nm_id,
            row.brand,
            row.seller_name_raw,
            row.wb_candidate_url,
            row.parse_status,
            row.parse_note,
        ])
    workbook.save(output_path)


def read_research_row(input_path: Path, row_number: int) -> ResearchRow:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [str(value).strip() if value is not None else "" for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    target = list(sheet.iter_rows(min_row=row_number, max_row=row_number, values_only=True))
    if not target:
        raise ValueError(f"Строка {row_number} не найдена")
    data = dict(zip(headers, target[0], strict=False))
    return ResearchRow(**data)


def read_research_rows_range(input_path: Path, start_row: int = 2, limit: int = 10) -> list[ResearchRow]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [str(value).strip() if value is not None else "" for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]

    result: list[ResearchRow] = []
    max_row = start_row + limit - 1
    for row_number in range(start_row, max_row + 1):
        target = list(sheet.iter_rows(min_row=row_number, max_row=row_number, values_only=True))
        if not target:
            continue
        data = dict(zip(headers, target[0], strict=False))
        result.append(ResearchRow(**data))
    return result


def save_batch_results(output_path: Path, rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "batch_results"
    headers = [
        "row_number",
        "source_sheet",
        "source_row_index",
        "product_name",
        "wb_nm_id",
        "brand",
        "seller_name_raw",
        "wb_candidate_url",
        "final_url",
        "seller_url",
        "navigated_to_seller_page",
        "seller_display_name",
        "entity_type",
        "inn",
        "ogrn",
        "ogrnip",
        "parse_status",
        "parse_note",
        "http_status",
        "used_persistent_profile",
        "screenshot_path",
        "html_path",
        "text_path",
        "marketguru_source_sheet",
        "marketguru_source_row_index",
        "marketguru_product_name",
        "marketguru_brand",
        "marketguru_seller_name",
        "marketguru_wb_nm_id",
        "marketguru_candidate_url",
        "wb_seller_name",
        "wb_seller_url",
    ]
    sheet.append(headers)
    for row in rows:
        normalized_row = dict(row)
        normalized_row.setdefault("marketguru_source_sheet", row.get("source_sheet"))
        normalized_row.setdefault("marketguru_source_row_index", row.get("source_row_index"))
        normalized_row.setdefault("marketguru_product_name", row.get("product_name"))
        normalized_row.setdefault("marketguru_brand", row.get("brand"))
        normalized_row.setdefault("marketguru_seller_name", row.get("seller_name_raw"))
        normalized_row.setdefault("marketguru_wb_nm_id", row.get("wb_nm_id"))
        normalized_row.setdefault("marketguru_candidate_url", row.get("wb_candidate_url"))
        normalized_row.setdefault("wb_seller_name", row.get("seller_display_name"))
        normalized_row.setdefault("wb_seller_url", row.get("seller_url"))
        sheet.append([normalized_row.get(header) for header in headers])
    workbook.save(output_path)


def merge_batch_results_with_compass(
    batch_results_path: Path,
    compass_path: Path,
    output_path: Path,
) -> dict[str, Any]:
    batch_rows = _read_sheet_as_dicts(batch_results_path)
    compass_payload = _read_compass_rows(compass_path)
    compass_index = compass_payload["index"]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "final_enriched"

    batch_headers = batch_payload_headers = batch_rows[0].keys() if batch_rows else []
    compass_headers = compass_payload["headers"]
    output_headers = list(batch_headers) + [
        "merge_inn_normalized",
        "compass_match_found",
        "compass_match_count",
        "compass_source_sheet",
        "compass_inn_header",
    ] + [f"compass_{header}" for header in compass_headers]
    sheet.append(output_headers)

    matched_rows = 0
    unmatched_rows = 0
    for row in batch_rows:
        normalized_inn = _normalize_inn_value(row.get("inn"))
        matches = compass_index.get(normalized_inn, []) if normalized_inn else []
        selected_match = matches[0] if matches else None
        if matches:
            matched_rows += 1
        else:
            unmatched_rows += 1

        output_row = dict(row)
        output_row["merge_inn_normalized"] = normalized_inn
        output_row["compass_match_found"] = bool(matches)
        output_row["compass_match_count"] = len(matches)
        output_row["compass_source_sheet"] = compass_payload["sheet_name"] if matches else None
        output_row["compass_inn_header"] = compass_payload["inn_header"] if matches else None
        for header in compass_headers:
            output_row[f"compass_{header}"] = selected_match.get(header) if selected_match else None
        sheet.append([output_row.get(header) for header in output_headers])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return {
        "batch_rows": len(batch_rows),
        "matched_rows": matched_rows,
        "unmatched_rows": unmatched_rows,
        "compass_sheet_name": compass_payload["sheet_name"],
        "compass_inn_header": compass_payload["inn_header"],
        "compass_rows_indexed": compass_payload["rows_indexed"],
        "output_path": str(output_path),
    }


def _read_compass_rows(compass_path: Path) -> dict[str, Any]:
    workbook = load_workbook(compass_path, read_only=True, data_only=True)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            continue
        headers = [_coerce_header(value) for value in header_row]
        if not any(headers):
            continue
        inn_header = _resolve_compass_inn_header(headers)
        if inn_header is None:
            continue

        index: dict[str, list[dict[str, Any]]] = {}
        rows_indexed = 0
        for row in rows_iter:
            row_dict = dict(zip(headers, row, strict=False))
            normalized_inn = _normalize_inn_value(row_dict.get(inn_header))
            if not normalized_inn:
                continue
            index.setdefault(normalized_inn, []).append(row_dict)
            rows_indexed += 1

        return {
            "sheet_name": sheet_name,
            "headers": headers,
            "inn_header": inn_header,
            "rows_indexed": rows_indexed,
            "index": index,
        }
    raise ValueError("В выгрузке Compass не найдена колонка ИНН. Проверь Excel и заголовки столбцов.")


def _read_sheet_as_dicts(input_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise ValueError(f"Пустой Excel: {input_path}") from exc
    headers = [_coerce_header(value) for value in header_row]
    result: list[dict[str, Any]] = []
    for row in rows_iter:
        row_dict = dict(zip(headers, row, strict=False))
        if _is_effectively_empty_row(row_dict.values()):
            continue
        result.append(row_dict)
    return result


def _resolve_compass_inn_header(headers: list[str]) -> str | None:
    normalized_map = {header.casefold(): header for header in headers if header}
    for candidate in COMPASS_INN_HEADER_CANDIDATES:
        found = normalized_map.get(candidate.casefold())
        if found:
            return found
    for header in headers:
        normalized = header.casefold()
        if "инн" in normalized:
            return header
    return None


def _normalize_inn_value(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float):
        if value.is_integer():
            value = int(value)
        else:
            value = str(value)
    text = str(value).strip()
    if not text:
        return None
    if text.endswith('.0'):
        text = text[:-2]
    digits = ''.join(ch for ch in text if ch.isdigit())
    return digits or None


def _normalize_key_part(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().casefold()


def _get_cell(row: list[Any] | tuple[Any, ...], header_index: dict[str, int], header_name: str) -> Any:
    index = header_index.get(header_name)
    if index is None or index >= len(row):
        return None
    return row[index]


def _normalize_sheet_selection(all_sheet_names: list[str], selected_sheets: list[str] | None) -> list[str]:
    if not selected_sheets:
        return list(all_sheet_names)

    missing = [sheet_name for sheet_name in selected_sheets if sheet_name not in all_sheet_names]
    if missing:
        raise ValueError(f"Не найдены листы: {', '.join(missing)}")
    return selected_sheets


def _analyze_sheet(sheet: Worksheet) -> SheetAnalyzeSummary:
    sheet_meta = _prepare_sheet_meta(sheet)
    if sheet_meta is None:
        return SheetAnalyzeSummary(
            sheet_name=sheet.title,
            status="SKIPPED",
            reason="На листе не найдены обязательные колонки: Товар / Артикул / Бренд / Продавец",
        )

    preview_rows = _collect_preview_rows(
        sheet,
        header_row_index=sheet_meta["header_row_index"],
        headers=sheet_meta["headers"],
    )

    return SheetAnalyzeSummary(
        sheet_name=sheet.title,
        status="VALID",
        detected_header_row=sheet_meta["header_row_index"],
        total_rows_scanned=sheet_meta["total_rows_scanned"],
        total_columns=len(sheet_meta["headers"]),
        data_rows_count=sheet_meta["data_rows_count"],
        headers=sheet_meta["headers"],
        preview_rows=preview_rows,
        has_seller_name=REQUIRED_HEADERS["seller"] in sheet_meta["header_index"],
        has_brand=REQUIRED_HEADERS["brand"] in sheet_meta["header_index"],
        has_wb_url=any(
            "wildberries" in header.lower() or "url" in header.lower() or "ссылка" in header.lower()
            for header in sheet_meta["headers"]
        ),
        has_article=REQUIRED_HEADERS["article"] in sheet_meta["header_index"],
        candidate_key_field=(
            REQUIRED_HEADERS["article"]
            if REQUIRED_HEADERS["article"] in sheet_meta["header_index"]
            else REQUIRED_HEADERS["seller"]
        ),
    )


def _prepare_sheet_meta(sheet: Worksheet) -> dict[str, Any] | None:
    scanned_rows: list[list[Any]] = []
    header_row_index: int | None = None
    headers: list[str] = []

    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        row_values = list(row)
        scanned_rows.append(row_values)
        candidate_headers = [_coerce_header(value) for value in row_values]
        if _is_required_header_row(candidate_headers):
            header_row_index = row_number
            headers = candidate_headers
            break
        if row_number >= HEADER_SCAN_LIMIT:
            break

    if header_row_index is None:
        return None

    header_index = {header: idx for idx, header in enumerate(headers) if header}
    data_rows_count = 0
    for row in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
        if _is_effectively_empty_row(row):
            continue
        data_rows_count += 1

    return {
        "header_row_index": header_row_index,
        "headers": headers,
        "header_index": header_index,
        "total_rows_scanned": len(scanned_rows),
        "data_rows_count": data_rows_count,
    }


def _collect_preview_rows(sheet: Worksheet, header_row_index: int, headers: list[str]) -> list[dict[str, Any]]:
    preview_rows: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
        row_values = list(row)
        if _is_effectively_empty_row(row_values):
            continue
        preview_rows.append(dict(zip(headers, row_values, strict=False)))
        if len(preview_rows) >= PREVIEW_LIMIT:
            break
    return preview_rows


def _is_required_header_row(headers: Iterable[str]) -> bool:
    header_set = {header.strip() for header in headers if header and header.strip()}
    return all(required_header in header_set for required_header in REQUIRED_HEADERS.values())


def _coerce_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _coerce_to_string(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        normalized = value.strip()
        return normalized or None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    normalized = str(value).strip()
    return normalized or None


def _coerce_wb_nm_id(value: Any) -> int | None:
    if value in (None, "", "—"):
        return None
    try:
        if isinstance(value, float):
            if not value.is_integer():
                return None
            return int(value)
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def _is_effectively_empty_row(row: Iterable[Any]) -> bool:
    for value in row:
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return False
    return True



def export_unique_inn_list(batch_results_path: Path, output_path: Path) -> dict[str, Any]:
    batch_rows = _read_sheet_as_dicts(batch_results_path)
    seen: set[str] = set()
    output_rows: list[dict[str, Any]] = []
    for row in batch_rows:
        normalized_inn = _normalize_inn_value(row.get("inn"))
        if not normalized_inn or normalized_inn in seen:
            continue
        seen.add(normalized_inn)
        output_rows.append({
            "inn": normalized_inn,
            "marketguru_seller_name": row.get("marketguru_seller_name") or row.get("seller_name_raw"),
            "marketguru_brand": row.get("marketguru_brand") or row.get("brand"),
            "marketguru_product_name": row.get("marketguru_product_name") or row.get("product_name"),
            "marketguru_source_sheet": row.get("marketguru_source_sheet") or row.get("source_sheet"),
            "wb_seller_name": row.get("wb_seller_name") or row.get("seller_display_name"),
            "wb_seller_url": row.get("wb_seller_url") or row.get("seller_url"),
        })

    output_path.parent.mkdir(parents=True, exist_ok=True)
    _write_dict_rows_to_excel(output_path, output_rows, preferred_headers=[
        "inn",
        "marketguru_seller_name",
        "marketguru_brand",
        "marketguru_product_name",
        "marketguru_source_sheet",
        "wb_seller_name",
        "wb_seller_url",
    ], sheet_name="inn_for_compass")
    return {
        "batch_rows": len(batch_rows),
        "inn_rows": len(output_rows),
        "output_path": str(output_path),
    }



def export_batch_no_inn(batch_results_path: Path, output_path: Path) -> dict[str, Any]:
    batch_rows = _read_sheet_as_dicts(batch_results_path)
    output_rows = [row for row in batch_rows if not _normalize_inn_value(row.get("inn"))]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    _write_dict_rows_to_excel(output_path, output_rows, sheet_name="batch_no_inn")
    return {
        "batch_rows": len(batch_rows),
        "no_inn_rows": len(output_rows),
        "output_path": str(output_path),
    }



def export_compass_unmatched(final_enriched_path: Path, output_path: Path) -> dict[str, Any]:
    enriched_rows = _read_sheet_as_dicts(final_enriched_path)
    output_rows = [row for row in enriched_rows if not _coerce_to_bool(row.get("compass_match_found"))]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    _write_dict_rows_to_excel(output_path, output_rows, sheet_name="compass_unmatched")
    return {
        "final_rows": len(enriched_rows),
        "unmatched_rows": len(output_rows),
        "output_path": str(output_path),
    }


def update_inn_registry_from_batch_files(
    batch_results_paths: list[Path],
    registry_path: Path,
    new_inn_output_path: Path | None = None,
) -> dict[str, Any]:
    imported_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    registry_rows, import_rows = _read_inn_registry_payload(registry_path)
    registry_by_inn: dict[str, dict[str, Any]] = {}
    for row in registry_rows:
        normalized_inn = _normalize_inn_value(row.get("inn"))
        if normalized_inn:
            row["inn"] = normalized_inn
            registry_by_inn.setdefault(normalized_inn, row)

    total_rows = 0
    rows_with_inn = 0
    skipped_without_inn = 0
    already_known = 0
    duplicate_in_import = 0
    new_rows: list[dict[str, Any]] = []
    seen_in_import: set[str] = set()
    per_file_summaries: list[dict[str, Any]] = []

    for batch_path in batch_results_paths:
        _validate_batch_results_file(batch_path)
        batch_rows = _read_sheet_as_dicts(batch_path)
        category = _infer_category_from_batch_path(batch_path)
        file_rows_total = len(batch_rows)
        file_rows_with_inn = 0
        file_new = 0
        file_known = 0
        file_duplicates = 0
        file_skipped_without_inn = 0
        seen_in_file: set[str] = set()

        for row in batch_rows:
            total_rows += 1
            normalized_inn = _normalize_inn_value(row.get("inn"))
            if not normalized_inn:
                skipped_without_inn += 1
                file_skipped_without_inn += 1
                continue

            rows_with_inn += 1
            file_rows_with_inn += 1

            if normalized_inn in seen_in_file:
                duplicate_in_import += 1
                file_duplicates += 1
                continue
            seen_in_file.add(normalized_inn)

            if normalized_inn in registry_by_inn:
                already_known += 1
                file_known += 1
                existing = registry_by_inn[normalized_inn]
                existing["seen_count"] = _coerce_int(existing.get("seen_count"), default=1) + 1
                existing["last_seen_at"] = imported_at
                existing["last_source_batch"] = str(batch_path)
                continue

            if normalized_inn in seen_in_import:
                duplicate_in_import += 1
                file_duplicates += 1
                continue
            seen_in_import.add(normalized_inn)

            registry_row = _build_registry_row(
                inn=normalized_inn,
                batch_row=row,
                batch_path=batch_path,
                category=category,
                imported_at=imported_at,
            )
            registry_rows.append(registry_row)
            registry_by_inn[normalized_inn] = registry_row
            new_rows.append(registry_row)
            file_new += 1

        per_file_summaries.append({
            "source_batch_file": str(batch_path),
            "category": category,
            "rows_total": file_rows_total,
            "rows_with_inn": file_rows_with_inn,
            "unique_inn_in_file": len(seen_in_file),
            "new_inn_added": file_new,
            "already_known": file_known,
            "duplicate_in_file": file_duplicates,
            "skipped_without_inn": file_skipped_without_inn,
        })
        import_rows.append({
            "imported_at": imported_at,
            "source_batch_file": str(batch_path),
            "category": category,
            "rows_total": file_rows_total,
            "rows_with_inn": file_rows_with_inn,
            "unique_inn_in_file": len(seen_in_file),
            "new_inn_added": file_new,
            "already_known": file_known,
            "duplicate_in_file": file_duplicates,
            "skipped_without_inn": file_skipped_without_inn,
        })

    registry_path.parent.mkdir(parents=True, exist_ok=True)
    _write_inn_registry_workbook(registry_path, registry_rows, import_rows)

    if new_inn_output_path is not None:
        new_inn_output_path.parent.mkdir(parents=True, exist_ok=True)
        _write_dict_rows_to_excel(
            new_inn_output_path,
            [{"inn": row.get("inn")} for row in new_rows],
            preferred_headers=["inn"],
            sheet_name="inn_for_kontur",
        )

    return {
        "batch_files": len(batch_results_paths),
        "rows_total": total_rows,
        "rows_with_inn": rows_with_inn,
        "registry_rows": len(registry_by_inn),
        "new_inn_added": len(new_rows),
        "already_known": already_known,
        "duplicate_in_import": duplicate_in_import,
        "skipped_without_inn": skipped_without_inn,
        "registry_path": str(registry_path),
        "new_inn_output_path": str(new_inn_output_path) if new_inn_output_path else None,
        "files": per_file_summaries,
    }


def update_seller_history_from_batch_files(
    batch_results_paths: list[Path],
    history_path: Path,
) -> dict[str, Any]:
    imported_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    history_rows, import_rows = _read_seller_history_payload(history_path)
    history_by_key: dict[str, dict[str, Any]] = {}
    for row in history_rows:
        normalized_row = _normalize_seller_history_row(row)
        seller_key = normalized_row.get("seller_key")
        if seller_key:
            history_by_key.setdefault(seller_key, normalized_row)

    total_rows = 0
    rows_with_seller = 0
    skipped_without_seller = 0
    already_known = 0
    duplicate_in_import = 0
    skip_recommended_total = 0
    new_rows: list[dict[str, Any]] = []
    per_file_summaries: list[dict[str, Any]] = []

    for batch_path in batch_results_paths:
        _validate_batch_results_file(batch_path)
        batch_rows = _read_sheet_as_dicts(batch_path)
        category = _infer_category_from_batch_path(batch_path)
        file_rows_total = len(batch_rows)
        file_rows_with_seller = 0
        file_new = 0
        file_known = 0
        file_duplicates = 0
        file_skipped_without_seller = 0
        file_skip_recommended = 0
        seen_in_file: set[str] = set()

        for row in batch_rows:
            total_rows += 1
            seller_key = _extract_batch_seller_key(row)
            if not seller_key:
                skipped_without_seller += 1
                file_skipped_without_seller += 1
                continue

            rows_with_seller += 1
            file_rows_with_seller += 1

            if seller_key in seen_in_file:
                duplicate_in_import += 1
                file_duplicates += 1
                continue
            seen_in_file.add(seller_key)

            skip_recommended = _is_seller_history_skip_candidate(row)
            if skip_recommended:
                skip_recommended_total += 1
                file_skip_recommended += 1

            existing = history_by_key.get(seller_key)
            if existing is not None:
                already_known += 1
                file_known += 1
                _merge_seller_history_row(
                    existing=existing,
                    source_row=row,
                    source_path=batch_path,
                    category=category,
                    imported_at=imported_at,
                    skip_recommended=skip_recommended,
                )
                continue

            history_row = _build_seller_history_row(
                seller_key=seller_key,
                source_row=row,
                source_path=batch_path,
                category=category,
                imported_at=imported_at,
                skip_recommended=skip_recommended,
            )
            history_rows.append(history_row)
            history_by_key[seller_key] = history_row
            new_rows.append(history_row)
            file_new += 1

        file_summary = {
            "source_batch_file": str(batch_path),
            "category": category,
            "rows_total": file_rows_total,
            "rows_with_seller": file_rows_with_seller,
            "unique_seller_in_file": len(seen_in_file),
            "new_seller_added": file_new,
            "already_known": file_known,
            "duplicate_in_file": file_duplicates,
            "skipped_without_seller": file_skipped_without_seller,
            "skip_recommended_in_file": file_skip_recommended,
        }
        per_file_summaries.append(file_summary)
        import_rows.append({
            "imported_at": imported_at,
            **file_summary,
        })

    history_path.parent.mkdir(parents=True, exist_ok=True)
    _write_seller_history_workbook(history_path, history_rows, import_rows)

    return {
        "batch_files": len(batch_results_paths),
        "rows_total": total_rows,
        "rows_with_seller": rows_with_seller,
        "history_rows": len(history_by_key),
        "new_seller_added": len(new_rows),
        "already_known": already_known,
        "duplicate_in_import": duplicate_in_import,
        "skipped_without_seller": skipped_without_seller,
        "skip_recommended_total": skip_recommended_total,
        "history_path": str(history_path),
        "files": per_file_summaries,
    }


def import_seller_history_from_registry_files(
    registry_paths: list[Path],
    history_path: Path,
) -> dict[str, Any]:
    imported_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    history_rows, import_rows = _read_seller_history_payload(history_path)
    history_by_key: dict[str, dict[str, Any]] = {}
    for row in history_rows:
        normalized_row = _normalize_seller_history_row(row)
        seller_key = normalized_row.get("seller_key")
        if seller_key:
            history_by_key.setdefault(seller_key, normalized_row)

    total_rows = 0
    rows_with_seller = 0
    skipped_without_seller = 0
    already_known = 0
    duplicate_in_import = 0
    skip_recommended_total = 0
    new_rows: list[dict[str, Any]] = []
    per_file_summaries: list[dict[str, Any]] = []

    for registry_path in registry_paths:
        if not registry_path.exists():
            raise FileNotFoundError(f"Не найден файл для импорта в историю sellers: {registry_path}")

        source_rows, source_sheet_name = _read_seller_history_source_rows_from_registry_workbook(registry_path)
        file_rows_total = len(source_rows)
        file_rows_with_seller = 0
        file_new = 0
        file_known = 0
        file_duplicates = 0
        file_skipped_without_seller = 0
        file_skip_recommended = 0
        seen_in_file: set[str] = set()

        for row in source_rows:
            total_rows += 1
            seller_key = _extract_batch_seller_key(row)
            if not seller_key:
                skipped_without_seller += 1
                file_skipped_without_seller += 1
                continue

            rows_with_seller += 1
            file_rows_with_seller += 1

            if seller_key in seen_in_file:
                duplicate_in_import += 1
                file_duplicates += 1
                continue
            seen_in_file.add(seller_key)

            row_category = _coerce_to_string(row.get("first_category")) or _coerce_to_string(row.get("last_category")) or source_sheet_name
            skip_recommended = _is_seller_history_skip_candidate(row)
            if skip_recommended:
                skip_recommended_total += 1
                file_skip_recommended += 1

            existing = history_by_key.get(seller_key)
            if existing is not None:
                already_known += 1
                file_known += 1
                _merge_seller_history_row(
                    existing=existing,
                    source_row=row,
                    source_path=registry_path,
                    category=row_category,
                    imported_at=imported_at,
                    skip_recommended=skip_recommended,
                )
                continue

            history_row = _build_seller_history_row(
                seller_key=seller_key,
                source_row=row,
                source_path=registry_path,
                category=row_category,
                imported_at=imported_at,
                skip_recommended=skip_recommended,
            )
            history_rows.append(history_row)
            history_by_key[seller_key] = history_row
            new_rows.append(history_row)
            file_new += 1

        file_summary = {
            "source_batch_file": str(registry_path),
            "category": source_sheet_name,
            "rows_total": file_rows_total,
            "rows_with_seller": file_rows_with_seller,
            "unique_seller_in_file": len(seen_in_file),
            "new_seller_added": file_new,
            "already_known": file_known,
            "duplicate_in_file": file_duplicates,
            "skipped_without_seller": file_skipped_without_seller,
            "skip_recommended_in_file": file_skip_recommended,
        }
        per_file_summaries.append(file_summary)
        import_rows.append({
            "imported_at": imported_at,
            **file_summary,
        })

    history_path.parent.mkdir(parents=True, exist_ok=True)
    _write_seller_history_workbook(history_path, history_rows, import_rows)

    return {
        "source_files": len(registry_paths),
        "rows_total": total_rows,
        "rows_with_seller": rows_with_seller,
        "history_rows": len(history_by_key),
        "new_seller_added": len(new_rows),
        "already_known": already_known,
        "duplicate_in_import": duplicate_in_import,
        "skipped_without_seller": skipped_without_seller,
        "skip_recommended_total": skip_recommended_total,
        "history_path": str(history_path),
        "files": per_file_summaries,
    }


def merge_inn_registry_files(
    primary_registry_path: Path,
    secondary_registry_path: Path,
    output_path: Path,
) -> dict[str, Any]:
    if not primary_registry_path.exists():
        raise FileNotFoundError(f"Не найден главный реестр: {primary_registry_path}")
    if not secondary_registry_path.exists():
        raise FileNotFoundError(f"Не найден второй реестр: {secondary_registry_path}")
    if primary_registry_path.resolve() == secondary_registry_path.resolve():
        raise ValueError("Для объединения нужны два разных файла реестра")

    primary_rows, primary_import_rows = _read_inn_registry_payload(primary_registry_path)
    secondary_rows, secondary_import_rows = _read_inn_registry_payload(secondary_registry_path)

    merged_rows: list[dict[str, Any]] = []
    merged_by_inn: dict[str, dict[str, Any]] = {}

    primary_unique_inn = 0
    primary_duplicates = 0
    primary_skipped_without_inn = 0

    for row in primary_rows:
        normalized_row = _normalize_registry_row(row)
        normalized_inn = normalized_row.get("inn")
        if not normalized_inn:
            primary_skipped_without_inn += 1
            continue
        if normalized_inn in merged_by_inn:
            primary_duplicates += 1
            continue
        merged_rows.append(normalized_row)
        merged_by_inn[normalized_inn] = normalized_row
        primary_unique_inn += 1

    secondary_unique_seen: set[str] = set()
    secondary_duplicates = 0
    secondary_skipped_without_inn = 0
    overlaps_with_primary = 0
    added_from_secondary = 0

    for row in secondary_rows:
        normalized_row = _normalize_registry_row(row)
        normalized_inn = normalized_row.get("inn")
        if not normalized_inn:
            secondary_skipped_without_inn += 1
            continue
        if normalized_inn in secondary_unique_seen:
            secondary_duplicates += 1
            continue
        secondary_unique_seen.add(normalized_inn)

        if normalized_inn in merged_by_inn:
            overlaps_with_primary += 1
            continue

        merged_rows.append(normalized_row)
        merged_by_inn[normalized_inn] = normalized_row
        added_from_secondary += 1

    merged_import_rows = [
        _normalize_import_row(row) for row in primary_import_rows if not _is_effectively_empty_row(row.values())
    ]
    merged_import_rows.extend(
        _normalize_import_row(row) for row in secondary_import_rows if not _is_effectively_empty_row(row.values())
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    _write_inn_registry_workbook(output_path, merged_rows, merged_import_rows)

    return {
        "primary_registry_path": str(primary_registry_path),
        "secondary_registry_path": str(secondary_registry_path),
        "output_path": str(output_path),
        "primary_rows_total": len(primary_rows),
        "secondary_rows_total": len(secondary_rows),
        "primary_unique_inn": primary_unique_inn,
        "secondary_unique_inn": len(secondary_unique_seen),
        "primary_duplicates": primary_duplicates,
        "secondary_duplicates": secondary_duplicates,
        "primary_skipped_without_inn": primary_skipped_without_inn,
        "secondary_skipped_without_inn": secondary_skipped_without_inn,
        "overlaps_with_primary": overlaps_with_primary,
        "added_from_secondary": added_from_secondary,
        "merged_registry_rows": len(merged_rows),
        "imports_rows": len(merged_import_rows),
    }


def discover_batch_results(root_dir: Path) -> list[Path]:
    if root_dir.is_file():
        return [root_dir] if _is_batch_results_filename(root_dir) else []
    if not root_dir.exists():
        raise FileNotFoundError(f"Папка не найдена: {root_dir}")
    return sorted(
        (path for path in root_dir.rglob("*.xlsx") if _is_batch_results_filename(path)),
        key=lambda path: str(path).casefold(),
    )


def _is_batch_results_filename(path: Path) -> bool:
    return path.name.casefold() == "batch_results.xlsx"


def _validate_batch_results_file(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(f"Не найден batch_results.xlsx: {path}")
    if not _is_batch_results_filename(path):
        raise ValueError(f"В реестр можно импортировать только файлы batch_results.xlsx: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise ValueError(f"Пустой batch_results.xlsx: {path}") from exc
    headers = {_coerce_header(value) for value in header_row if _coerce_header(value)}
    missing = sorted(BATCH_RESULTS_REQUIRED_HEADERS - headers)
    if missing:
        raise ValueError(
            f"Файл не похож на batch_results.xlsx: {path}. Не найдены колонки: {', '.join(missing)}"
        )


def _read_inn_registry_payload(registry_path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    if not registry_path.exists():
        return [], []

    workbook = load_workbook(registry_path, read_only=True, data_only=True)
    registry_sheet_name = "inn_registry" if "inn_registry" in workbook.sheetnames else workbook.sheetnames[0]
    registry_rows = _read_worksheet_as_dicts(workbook[registry_sheet_name])
    import_rows = _read_worksheet_as_dicts(workbook["imports"]) if "imports" in workbook.sheetnames else []
    return registry_rows, import_rows


def _read_seller_history_payload(history_path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    if not history_path.exists():
        return [], []

    workbook = load_workbook(history_path, read_only=True, data_only=True)
    history_sheet_name = "seller_history" if "seller_history" in workbook.sheetnames else workbook.sheetnames[0]
    history_rows = _read_worksheet_as_dicts(workbook[history_sheet_name])
    import_rows = _read_worksheet_as_dicts(workbook["imports"]) if "imports" in workbook.sheetnames else []
    return history_rows, import_rows


def _read_worksheet_as_dicts(sheet: Worksheet) -> list[dict[str, Any]]:
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    headers = [_coerce_header(value) for value in header_row]
    result: list[dict[str, Any]] = []
    for row in rows_iter:
        row_dict = dict(zip(headers, row, strict=False))
        if _is_effectively_empty_row(row_dict.values()):
            continue
        result.append(row_dict)
    return result


def load_known_seller_keys_from_registry(registry_path: Path) -> set[str]:
    return set(load_registry_seller_index(registry_path))


def load_registry_seller_index(registry_path: Path) -> dict[str, dict[str, Any]]:
    registry_rows, _ = _read_inn_registry_payload(registry_path)
    seller_index: dict[str, dict[str, Any]] = {}
    for row in registry_rows:
        normalized_row = _normalize_registry_row(row)
        for field_name in ("seller_name_raw", "seller_display_name"):
            seller_key = _normalize_key_part(normalized_row.get(field_name))
            if seller_key:
                seller_index.setdefault(seller_key, normalized_row)
    return seller_index


def load_skip_seller_index_from_history(history_path: Path) -> dict[str, dict[str, Any]]:
    history_rows, _ = _read_seller_history_payload(history_path)
    seller_index: dict[str, dict[str, Any]] = {}
    for row in history_rows:
        normalized_row = _normalize_seller_history_row(row)
        seller_key = normalized_row.get("seller_key")
        if not seller_key or not _coerce_to_bool(normalized_row.get("skip_recommended")):
            continue
        seller_index.setdefault(seller_key, normalized_row)
    return seller_index


def load_known_seller_sources(
    registry_path: Path | None = None,
    seller_history_path: Path | None = None,
) -> dict[str, dict[str, Any]]:
    known_sources: dict[str, dict[str, Any]] = {}

    if seller_history_path is not None and seller_history_path.exists():
        for seller_key, row in load_skip_seller_index_from_history(seller_history_path).items():
            known_sources.setdefault(seller_key, {
                "source": "seller_history",
                "path": seller_history_path,
                "row": row,
            })

    if registry_path is not None and registry_path.exists():
        for seller_key, row in load_registry_seller_index(registry_path).items():
            known_sources[seller_key] = {
                "source": "inn_registry",
                "path": registry_path,
                "row": row,
            }

    return known_sources


def build_known_seller_batch_row(
    row_number: int,
    research_row: ResearchRow,
    registry_row: dict[str, Any],
    registry_path: Path,
) -> dict[str, Any]:
    seller_display_name = registry_row.get("seller_display_name") or research_row.seller_name_raw
    seller_url = registry_row.get("seller_url")
    return {
        "row_number": row_number,
        "source_sheet": research_row.source_sheet,
        "source_row_index": research_row.source_row_index,
        "product_name": research_row.product_name,
        "wb_nm_id": research_row.wb_nm_id,
        "brand": research_row.brand,
        "seller_name_raw": research_row.seller_name_raw,
        "wb_candidate_url": research_row.wb_candidate_url,
        "final_url": seller_url,
        "seller_url": seller_url,
        "navigated_to_seller_page": False,
        "seller_display_name": seller_display_name,
        "entity_type": registry_row.get("entity_type"),
        "inn": registry_row.get("inn"),
        "ogrn": registry_row.get("ogrn"),
        "ogrnip": registry_row.get("ogrnip"),
        "parse_status": "SKIPPED_KNOWN_SELLER",
        "parse_note": f"Продавец найден в реестре ИНН: {registry_path}",
        "http_status": None,
        "used_persistent_profile": False,
        "screenshot_path": None,
        "html_path": None,
        "text_path": None,
        "marketguru_source_sheet": research_row.source_sheet,
        "marketguru_source_row_index": research_row.source_row_index,
        "marketguru_product_name": research_row.product_name,
        "marketguru_brand": research_row.brand,
        "marketguru_seller_name": research_row.seller_name_raw,
        "marketguru_wb_nm_id": research_row.wb_nm_id,
        "marketguru_candidate_url": research_row.wb_candidate_url,
        "wb_seller_name": seller_display_name,
        "wb_seller_url": seller_url,
    }


def build_viewed_seller_batch_row(
    row_number: int,
    research_row: ResearchRow,
    history_row: dict[str, Any],
    history_path: Path,
) -> dict[str, Any]:
    seller_display_name = history_row.get("seller_display_name") or research_row.seller_name_raw
    seller_url = history_row.get("seller_url")
    last_status = history_row.get("last_parse_status")
    note_suffix = f" Последний статус: {last_status}." if last_status else ""
    return {
        "row_number": row_number,
        "source_sheet": research_row.source_sheet,
        "source_row_index": research_row.source_row_index,
        "product_name": research_row.product_name,
        "wb_nm_id": research_row.wb_nm_id,
        "brand": research_row.brand,
        "seller_name_raw": research_row.seller_name_raw,
        "wb_candidate_url": research_row.wb_candidate_url,
        "final_url": seller_url,
        "seller_url": seller_url,
        "navigated_to_seller_page": False,
        "seller_display_name": seller_display_name,
        "entity_type": history_row.get("entity_type"),
        "inn": history_row.get("inn"),
        "ogrn": history_row.get("ogrn"),
        "ogrnip": history_row.get("ogrnip"),
        "parse_status": "SKIPPED_VIEWED_SELLER",
        "parse_note": f"Продавец найден в истории просмотренных sellers: {history_path}.{note_suffix}".strip(),
        "http_status": None,
        "used_persistent_profile": False,
        "screenshot_path": None,
        "html_path": None,
        "text_path": None,
        "marketguru_source_sheet": research_row.source_sheet,
        "marketguru_source_row_index": research_row.source_row_index,
        "marketguru_product_name": research_row.product_name,
        "marketguru_brand": research_row.brand,
        "marketguru_seller_name": research_row.seller_name_raw,
        "marketguru_wb_nm_id": research_row.wb_nm_id,
        "marketguru_candidate_url": research_row.wb_candidate_url,
        "wb_seller_name": seller_display_name,
        "wb_seller_url": seller_url,
    }


def _build_registry_row(
    inn: str,
    batch_row: dict[str, Any],
    batch_path: Path,
    category: str,
    imported_at: str,
) -> dict[str, Any]:
    return {
        "inn": inn,
        "first_category": category,
        "first_source_batch": str(batch_path),
        "first_seen_at": imported_at,
        "seller_name_raw": batch_row.get("marketguru_seller_name") or batch_row.get("seller_name_raw"),
        "seller_display_name": batch_row.get("wb_seller_name") or batch_row.get("seller_display_name"),
        "entity_type": batch_row.get("entity_type"),
        "ogrn": batch_row.get("ogrn"),
        "ogrnip": batch_row.get("ogrnip"),
        "seller_url": batch_row.get("wb_seller_url") or batch_row.get("seller_url"),
        "wb_candidate_url": batch_row.get("marketguru_candidate_url") or batch_row.get("wb_candidate_url"),
        "parse_status": batch_row.get("parse_status"),
        "note": batch_row.get("parse_note") or batch_row.get("note"),
        "seen_count": 1,
        "last_seen_at": imported_at,
        "last_source_batch": str(batch_path),
    }


def _normalize_registry_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized_row = {header: row.get(header) for header in INN_REGISTRY_HEADERS}
    normalized_row["inn"] = _normalize_inn_value(normalized_row.get("inn"))
    normalized_row["seen_count"] = _coerce_int(normalized_row.get("seen_count"), default=1)
    return normalized_row


def _normalize_import_row(row: dict[str, Any]) -> dict[str, Any]:
    return {header: row.get(header) for header in INN_IMPORT_HEADERS}


def _extract_batch_seller_key(batch_row: dict[str, Any]) -> str:
    seller_value = (
        batch_row.get("marketguru_seller_name")
        or batch_row.get("seller_name_raw")
        or batch_row.get("wb_seller_name")
        or batch_row.get("first_seller_name_raw")
        or batch_row.get("last_seller_name_raw")
        or batch_row.get("seller_display_name")
    )
    return _normalize_key_part(seller_value)


def _is_seller_history_skip_candidate(batch_row: dict[str, Any]) -> bool:
    if _normalize_inn_value(batch_row.get("inn")):
        return True
    parse_status = _coerce_to_string(batch_row.get("parse_status"))
    return bool(parse_status and parse_status in SELLER_HISTORY_SKIP_STATUSES)


def _build_seller_history_row(
    seller_key: str,
    source_row: dict[str, Any],
    source_path: Path,
    category: str,
    imported_at: str,
    skip_recommended: bool,
) -> dict[str, Any]:
    seller_name_raw = (
        source_row.get("marketguru_seller_name")
        or source_row.get("seller_name_raw")
        or source_row.get("last_seller_name_raw")
        or source_row.get("first_seller_name_raw")
    )
    seller_display_name = source_row.get("wb_seller_name") or source_row.get("seller_display_name")
    seller_url = source_row.get("wb_seller_url") or source_row.get("seller_url")
    parse_status = (
        source_row.get("parse_status")
        or source_row.get("last_parse_status")
        or source_row.get("first_parse_status")
    )
    note = (
        source_row.get("parse_note")
        or source_row.get("note")
        or source_row.get("last_note")
        or source_row.get("first_note")
    )
    first_seen_at = source_row.get("first_seen_at") or imported_at
    last_seen_at = source_row.get("last_seen_at") or source_row.get("first_seen_at") or imported_at
    first_source_batch = source_row.get("first_source_batch") or str(source_path)
    last_source_batch = source_row.get("last_source_batch") or str(source_path)
    first_category = source_row.get("first_category") or category
    last_category = source_row.get("last_category") or source_row.get("first_category") or category
    seen_count = max(_coerce_int(source_row.get("seen_count"), default=1), 1)
    return {
        "seller_key": seller_key,
        "first_seller_name_raw": seller_name_raw,
        "last_seller_name_raw": seller_name_raw,
        "seller_display_name": seller_display_name,
        "seller_url": seller_url,
        "entity_type": source_row.get("entity_type"),
        "inn": _normalize_inn_value(source_row.get("inn")),
        "ogrn": source_row.get("ogrn"),
        "ogrnip": source_row.get("ogrnip"),
        "first_parse_status": parse_status,
        "last_parse_status": parse_status,
        "first_note": note,
        "last_note": note,
        "skip_recommended": skip_recommended,
        "seen_count": seen_count,
        "first_seen_at": first_seen_at,
        "last_seen_at": last_seen_at,
        "first_source_batch": first_source_batch,
        "last_source_batch": last_source_batch,
        "first_category": first_category,
        "last_category": last_category,
    }


def _merge_seller_history_row(
    existing: dict[str, Any],
    source_row: dict[str, Any],
    source_path: Path,
    category: str,
    imported_at: str,
    skip_recommended: bool,
) -> None:
    source_seen_count = max(_coerce_int(source_row.get("seen_count"), default=1), 1)
    seller_name_raw = (
        source_row.get("marketguru_seller_name")
        or source_row.get("seller_name_raw")
        or source_row.get("last_seller_name_raw")
        or source_row.get("first_seller_name_raw")
    )
    parse_status = (
        source_row.get("parse_status")
        or source_row.get("last_parse_status")
        or source_row.get("first_parse_status")
    )
    note = (
        source_row.get("parse_note")
        or source_row.get("note")
        or source_row.get("last_note")
        or source_row.get("first_note")
    )

    existing["seen_count"] = _coerce_int(existing.get("seen_count"), default=1) + source_seen_count
    existing["last_seen_at"] = source_row.get("last_seen_at") or source_row.get("first_seen_at") or imported_at
    existing["last_source_batch"] = source_row.get("last_source_batch") or str(source_path)
    existing["last_category"] = source_row.get("last_category") or source_row.get("first_category") or category
    if seller_name_raw:
        existing["last_seller_name_raw"] = seller_name_raw
        if not existing.get("first_seller_name_raw"):
            existing["first_seller_name_raw"] = seller_name_raw
    if source_row.get("wb_seller_name") or source_row.get("seller_display_name"):
        existing["seller_display_name"] = source_row.get("wb_seller_name") or source_row.get("seller_display_name")
    if source_row.get("wb_seller_url") or source_row.get("seller_url"):
        existing["seller_url"] = source_row.get("wb_seller_url") or source_row.get("seller_url")
    if source_row.get("entity_type"):
        existing["entity_type"] = source_row.get("entity_type")
    normalized_inn = _normalize_inn_value(source_row.get("inn"))
    if normalized_inn:
        existing["inn"] = normalized_inn
    if source_row.get("ogrn"):
        existing["ogrn"] = source_row.get("ogrn")
    if source_row.get("ogrnip"):
        existing["ogrnip"] = source_row.get("ogrnip")
    if parse_status:
        existing["last_parse_status"] = parse_status
        if not existing.get("first_parse_status"):
            existing["first_parse_status"] = parse_status
    if note:
        existing["last_note"] = note
        if not existing.get("first_note"):
            existing["first_note"] = note
    if not existing.get("first_seen_at"):
        existing["first_seen_at"] = source_row.get("first_seen_at") or imported_at
    if not existing.get("first_source_batch"):
        existing["first_source_batch"] = source_row.get("first_source_batch") or str(source_path)
    if not existing.get("first_category"):
        existing["first_category"] = source_row.get("first_category") or category
    existing["skip_recommended"] = _coerce_to_bool(existing.get("skip_recommended")) or skip_recommended


def _read_seller_history_source_rows_from_registry_workbook(registry_path: Path) -> tuple[list[dict[str, Any]], str]:
    workbook = load_workbook(registry_path, read_only=True, data_only=True)
    preferred_sheet_names = [
        sheet_name
        for sheet_name in ("inn_registry", "final_enriched")
        if sheet_name in workbook.sheetnames
    ]
    preferred_sheet_names.extend(
        sheet_name for sheet_name in workbook.sheetnames if sheet_name not in preferred_sheet_names
    )

    for sheet_name in preferred_sheet_names:
        rows = _read_worksheet_as_dicts(workbook[sheet_name])
        if not rows:
            continue
        headers = set(rows[0])
        if headers.intersection({"seller_name_raw", "marketguru_seller_name", "first_seller_name_raw", "last_seller_name_raw"}):
            return rows, sheet_name

    raise ValueError(
        f"В файле {registry_path} не найден лист с seller-данными для импорта в историю sellers"
    )


def _normalize_seller_history_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized_row = {header: row.get(header) for header in SELLER_HISTORY_HEADERS}
    normalized_row["seller_key"] = _normalize_key_part(normalized_row.get("seller_key"))
    normalized_row["inn"] = _normalize_inn_value(normalized_row.get("inn"))
    normalized_row["seen_count"] = _coerce_int(normalized_row.get("seen_count"), default=1)
    normalized_row["skip_recommended"] = _coerce_to_bool(normalized_row.get("skip_recommended"))
    return normalized_row


def _normalize_seller_history_import_row(row: dict[str, Any]) -> dict[str, Any]:
    return {header: row.get(header) for header in SELLER_HISTORY_IMPORT_HEADERS}


def _write_inn_registry_workbook(
    registry_path: Path,
    registry_rows: list[dict[str, Any]],
    import_rows: list[dict[str, Any]],
) -> None:
    workbook = Workbook()
    registry_sheet = workbook.active
    registry_sheet.title = "inn_registry"
    registry_sheet.append(INN_REGISTRY_HEADERS)
    for row in registry_rows:
        registry_sheet.append([row.get(header) for header in INN_REGISTRY_HEADERS])

    imports_sheet = workbook.create_sheet("imports")
    imports_sheet.append(INN_IMPORT_HEADERS)
    for row in import_rows:
        imports_sheet.append([row.get(header) for header in INN_IMPORT_HEADERS])

    _format_registry_sheet(registry_sheet)
    _format_imports_sheet(imports_sheet)
    workbook.save(registry_path)


def _write_seller_history_workbook(
    history_path: Path,
    history_rows: list[dict[str, Any]],
    import_rows: list[dict[str, Any]],
) -> None:
    workbook = Workbook()
    history_sheet = workbook.active
    history_sheet.title = "seller_history"
    history_sheet.append(SELLER_HISTORY_HEADERS)
    for row in history_rows:
        normalized_row = _normalize_seller_history_row(row)
        history_sheet.append([normalized_row.get(header) for header in SELLER_HISTORY_HEADERS])

    imports_sheet = workbook.create_sheet("imports")
    imports_sheet.append(SELLER_HISTORY_IMPORT_HEADERS)
    for row in import_rows:
        normalized_row = _normalize_seller_history_import_row(row)
        imports_sheet.append([normalized_row.get(header) for header in SELLER_HISTORY_IMPORT_HEADERS])

    _format_seller_history_sheet(history_sheet)
    _format_seller_history_imports_sheet(imports_sheet)
    workbook.save(history_path)


def _format_registry_sheet(sheet: Worksheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {
        "A": 28,
        "B": 28,
        "C": 28,
        "D": 14,
        "E": 16,
        "F": 16,
        "G": 18,
        "H": 42,
        "I": 44,
        "J": 18,
        "K": 42,
        "L": 12,
        "M": 20,
        "N": 20,
        "O": 65,
        "P": 65,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _format_seller_history_sheet(sheet: Worksheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {
        "A": 30,
        "B": 28,
        "C": 28,
        "D": 28,
        "E": 40,
        "F": 16,
        "G": 16,
        "H": 16,
        "I": 18,
        "J": 20,
        "K": 20,
        "L": 34,
        "M": 34,
        "N": 14,
        "O": 12,
        "P": 20,
        "Q": 20,
        "R": 60,
        "S": 60,
        "T": 28,
        "U": 28,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _format_seller_history_imports_sheet(sheet: Worksheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {
        "A": 20,
        "B": 65,
        "C": 28,
        "D": 12,
        "E": 14,
        "F": 16,
        "G": 16,
        "H": 16,
        "I": 14,
        "J": 18,
        "K": 18,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _format_imports_sheet(sheet: Worksheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {
        "A": 20,
        "B": 65,
        "C": 28,
        "D": 12,
        "E": 14,
        "F": 16,
        "G": 16,
        "H": 14,
        "I": 16,
        "J": 18,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _infer_category_from_batch_path(batch_path: Path) -> str:
    parent = batch_path.parent
    parts = list(parent.parts)
    lowered = [part.casefold() for part in parts]
    for marker in ("результаты", "results"):
        if marker in lowered:
            index = lowered.index(marker)
            tail = parts[index + 1 :]
            if tail:
                return " / ".join(tail)
    return parent.name


def _coerce_int(value: Any, default: int = 0) -> int:
    if value is None:
        return default
    try:
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return int(str(value).strip())
    except (TypeError, ValueError):
        return default



def _write_dict_rows_to_excel(
    output_path: Path,
    rows: list[dict[str, Any]],
    preferred_headers: list[str] | None = None,
    sheet_name: str = "Sheet1",
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    headers: list[str] = []
    if preferred_headers:
        headers.extend(preferred_headers)

    seen = set(headers)
    for row in rows:
        for header in row.keys():
            if header not in seen:
                headers.append(header)
                seen.add(header)

    if not headers:
        headers = preferred_headers or ["note"]
        sheet.append(headers)
        if headers == ["note"]:
            sheet.append(["Нет данных для сохранения"])
    else:
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header) for header in headers])

    workbook.save(output_path)



def _coerce_to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    text = str(value).strip().casefold()
    return text in {"1", "true", "yes", "да"}
